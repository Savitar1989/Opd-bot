
# file: opdtest_final.py
# NOTE: Replace BOT_TOKEN and WEBAPP_URL with your real values before running.
import asyncio
import os
import logging
import sqlite3
import json
import threading
import re
import requests
import urllib.parse
import math
import itertools
import time
from queue import Queue, Empty
from typing import Dict, List, Optional, Tuple

from flask import Flask, render_template_string, request, jsonify
from flask_cors import CORS

from flask import send_file
import openpyxl
from io import BytesIO

from collections import defaultdict


# Telegram imports are optional if you run bot; keep them to preserve original behavior
try:
    from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
    import urllib.parse
    from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
    TELEGRAM_AVAILABLE = True
except Exception:
    TELEGRAM_AVAILABLE = False

# =============== CONFIG ===============
BOT_TOKEN = "7741178469:AAH9pvClqBOa31Yenq_0Y9dxtrug-ZMmDk4"
WEBAPP_URL = "https://94377687755d.ngrok-free.app"
DB_NAME = "restaurant_orders.db"
ADMIN_USER_IDS = [7553912440]  # adjust as needed

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

notification_queue: "Queue[Dict]" = Queue()

# ---------------- Database Manager ----------------
class DatabaseManager:
    def __init__(self) -> None:
        self.init_db()
        self.db_path = DB_NAME
    def init_db(self) -> None:
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        # Összes tábla létrehozása egy kapcsolaton belül
        cur.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                restaurant_name TEXT NOT NULL,
                restaurant_address TEXT NOT NULL,
                phone_number TEXT,
                order_details TEXT NOT NULL,
                group_id INTEGER NOT NULL,
                group_name TEXT,
                message_id INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT (datetime('now', 'localtime')),
                status TEXT DEFAULT 'pending',
                delivery_partner_id INTEGER,
                delivery_partner_name TEXT,
                delivery_partner_username TEXT,
                estimated_time INTEGER,
                accepted_at TIMESTAMP,
                picked_up_at TIMESTAMP,
                delivered_at TIMESTAMP
            )
        """)
    
        cur.execute("""CREATE TABLE IF NOT EXISTS groups(id INTEGER PRIMARY KEY, name TEXT NOT NULL)""")
    
        cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_group ON orders(group_name)")
    
        # Futárok tábla létrehozása ugyanabban a kapcsolatban
        cur.execute("""CREATE TABLE IF NOT EXISTS couriers (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                first_name TEXT,
                last_name TEXT,
                last_seen TIMESTAMP DEFAULT (datetime('now', 'localtime'))
            )
        """)
    
        # Minden változtatás mentése és kapcsolat bezárása
        conn.commit()
        conn.close()

    def register_group(self, group_id: int, group_name: str) -> None:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("INSERT OR IGNORE INTO groups(id, name) VALUES (?,?)", (group_id, group_name))
        conn.commit(); conn.close()

    def save_order(self, item: Dict) -> int:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("""INSERT INTO orders (restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, message_id) VALUES (?,?,?,?,?,?,?)""",
                    (item.get("restaurant_name",""), item.get("restaurant_address",""), item.get("phone_number",""), item.get("order_details",""), item.get("group_id"), item.get("group_name"), item.get("message_id")))
        oid = cur.lastrowid; conn.commit(); conn.close(); return oid

    def get_open_orders(self) -> List[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, delivery_partner_id, estimated_time FROM orders WHERE status IN ('pending','accepted','picked_up') ORDER BY created_at DESC")
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return rows

    def get_order_by_id(self, order_id: int) -> Optional[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT * FROM orders WHERE id = ?", (order_id,)); row = cur.fetchone(); conn.close(); return dict(row) if row else None

    def update_order_status(self, order_id: int, status: str, partner_id: int | None = None, partner_name: str | None = None, partner_username: str | None = None, estimated_time: int | None = None) -> None:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("""
            UPDATE orders SET status = ?, delivery_partner_id = COALESCE(?, delivery_partner_id), delivery_partner_name = COALESCE(?, delivery_partner_name), delivery_partner_username = COALESCE(?, delivery_partner_username), estimated_time = COALESCE(?, estimated_time), accepted_at = CASE WHEN ?='accepted' THEN datetime('now', 'localtime') ELSE accepted_at END, picked_up_at = CASE WHEN ?='picked_up' THEN datetime ('now', 'localtime') ELSE picked_up_at END, delivered_at = CASE WHEN ?='delivered' THEN datetime('now', 'localtime') ELSE delivered_at END WHERE id = ?
        """, (status, partner_id, partner_name, partner_username, estimated_time, status, status, status, order_id))
        conn.commit(); conn.close()

    def get_partner_addresses(self, partner_id: int, status: str) -> List[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT id, restaurant_address, group_name FROM orders WHERE delivery_partner_id = ? AND status = ? ORDER BY created_at", (partner_id, status))
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return rows

    def get_partner_order_count(self, partner_id: int, status: str = None) -> int:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        if status: cur.execute("SELECT COUNT(*) FROM orders WHERE delivery_partner_id = ? AND status = ?", (partner_id, status))
        else: cur.execute("SELECT COUNT(*) FROM orders WHERE delivery_partner_id = ?", (partner_id,))
        c = cur.fetchone()[0]; conn.close(); return c

    def register_courier(self, user: Dict) -> None:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("""
            INSERT OR REPLACE INTO couriers(user_id, username, first_name, last_name, last_seen) 
            VALUES (?,?,?,?,datetime('now','localtime'))
        """, (user.get("id"), user.get("username"), user.get("first_name"), user.get("last_name")))
        conn.commit(); conn.close()

    def get_all_couriers(self) -> List[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT user_id, username, first_name, last_name FROM couriers")
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return rows

    def get_orders_by_courier(self, courier_id, status_filter=None):
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur = conn.cursor()

        if status_filter:
            cur.execute("""
                SELECT id, restaurant, address, phone, status, created_at
                FROM orders
                WHERE courier_id = ? AND status = ?
                ORDER BY created_at DESC
            """, (courier_id, status_filter))
        else:
            cur.execute("""
                SELECT id, restaurant, address, phone, status, created_at
                FROM orders
                WHERE courier_id = ?
                ORDER BY created_at DESC
            """, (courier_id,))

        rows = cur.fetchall()
        conn.close()
        return rows

db = DatabaseManager()

def notify_all_couriers_order(order_id: int, text: str):
    """
    Interaktív értesítés küldése inline keyboard-dal
    """
    try:
        from telegram import InlineKeyboardButton, InlineKeyboardMarkup
        
        # Gombok létrehozása
        keyboard = [
            [
                InlineKeyboardButton("⏱️ 10 perc", callback_data=f"accept_{order_id}_10"),
                InlineKeyboardButton("⏱️ 20 perc", callback_data=f"accept_{order_id}_20"),
                InlineKeyboardButton("⏱️ 30 perc", callback_data=f"accept_{order_id}_30")
            ],
            [InlineKeyboardButton("❌ Elutasítás", callback_data=f"reject_{order_id}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        couriers = db.get_all_couriers()
        logger.info(f"Found {len(couriers)} registered couriers")
        for c in couriers:
            uid = c.get("user_id")
            if uid:
                notification_queue.put({
                    "chat_id": uid, 
                    "text": text,
                    "reply_markup": reply_markup
                })
                logger.info(f"Queued interactive notification for courier {uid}")
    except Exception as e:
        logger.error(f"notify_all_couriers_order error: {e}")
1
        
# ---------------- Telegram Bot (kept intact) ----------------
class RestaurantBot:
    def __init__(self) -> None:
        if not TELEGRAM_AVAILABLE:
            logger.warning("python-telegram-bot not available; bot handlers won't be active in this environment.")
            return
        self.app = Application.builder().token(BOT_TOKEN).build()
        self._setup_handlers()

    def _setup_handlers(self) -> None:
        app = self.app
        app.add_handler(CommandHandler("start", self.start_cmd))
        app.add_handler(CommandHandler("help", self.help_cmd))
        app.add_handler(CommandHandler("register", self.register_group))
        app.add_handler(MessageHandler(filters.TEXT & filters.ChatType.GROUPS, self.handle_group_message))
        app.add_handler(CommandHandler("myorders", self.my_orders))
        app.add_handler(CallbackQueryHandler(self.handle_callback_query))
    
        if app.job_queue:
            app.job_queue.run_repeating(self.process_notifications, interval=3)

    async def process_notifications(self, context: ContextTypes.DEFAULT_TYPE):
        processed_count = 0; max_per_batch = 5
        while processed_count < max_per_batch:
            try: 
                item = notification_queue.get_nowait()
                processed_count += 1
                logger.info(f"Processing notification for chat_id: {item.get('chat_id')}")
            except Empty: 
                if processed_count == 0:
                    logger.debug("No notifications in queue")
                break
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    # Inline keyboard támogatás hozzáadása
                    kwargs = {
                        "chat_id": item["chat_id"],
                        "text": item.get("text", ""),
                        "parse_mode": "Markdown"
                    }
                    if "reply_markup" in item:
                        kwargs["reply_markup"] = item["reply_markup"]
                
                    await context.bot.send_message(**kwargs)
                    logger.info(f"Successfully sent notification to {item['chat_id']}")
                    break
                except Exception as e:
                    logger.error(f"Failed to send notification to {item['chat_id']} (attempt {attempt+1}): {e}")
                    if attempt < max_retries - 1: await asyncio.sleep(1)

    def send_notification(self, chat_id: int, text: str):
        try:
            if not text or not chat_id: return
            notification_queue.put({"chat_id": chat_id, "text": text})
        except Exception as e:
            logger.error(f"Error queueing notification: {e}")

    async def start_cmd(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if update.effective_chat.type == "private":
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🚚 Elérhető rendelések", web_app=WebAppInfo(url=f"{WEBAPP_URL}"))]])
            await update.message.reply_text(f"Üdv, {user.first_name}!\nNyisd meg a futár felületet:", reply_markup=kb)
        else:
            await update.message.reply_text("Használd a /register parancsot a csoport regisztrálásához.\nRendelés formátum:\nCím: ...\nTelefonszám: ...\nMegjegyzés: ...")

    async def help_cmd(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text("Rendelés formátum (csoportban):\n```\nCím: Budapest, Példa utca 1.\nTelefonszám: +36301234567\nMegjegyzés: kp / kártya / megjegyzés\n```", parse_mode="Markdown")

    async def register_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_chat.type not in ("group","supergroup"):
            await update.message.reply_text("Ezt a parancsot csoportban használd."); return
        gid = update.effective_chat.id; gname = update.effective_chat.title or "Ismeretlen csoport"
        db.register_group(gid, gname); await update.message.reply_text(f"✅ A '{gname}' csoport regisztrálva.")

    def parse_order_message(self, text: str) -> Dict | None:
        lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
        info = {}
        def after_colon(s: str) -> str:
            return s.split(":",1)[1].strip() if ":" in s else ""
        for ln in lines:
            low = ln.lower()
            if low.startswith("cím:") or low.startswith("cim:"): info["address"] = after_colon(ln)
            elif low.startswith("telefonszám:") or low.startswith("telefonszam:") or low.startswith("telefon:"): info["phone"] = after_colon(ln)
            elif low.startswith("megjegyzés:") or low.startswith("megjegyzes:"): info["details"] = after_colon(ln)
        if info.get("address"): info.setdefault("phone",""); info.setdefault("details",""); return info
        return None

    async def handle_group_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_chat.type not in ("group","supergroup"): return
        parsed = self.parse_order_message(update.message.text or ""); 
        if not parsed: return
        gid = update.effective_chat.id; gname = update.effective_chat.title or "Ismeretlen"
        item = {"restaurant_name": gname, "restaurant_address": parsed["address"], "phone_number": parsed.get("phone",""), "order_details": parsed.get("details",""), "group_id": gid, "group_name": gname, "message_id": update.message.message_id}
        order_id = db.save_order(item)
        await update.message.reply_text("✅ Rendelés rögzítve.\n\n" f"📍 Cím: {item['restaurant_address']}\n" f"📞 Telefon: {item['phone_number'] or '—'}\n" f"📝 Megjegyzés: {item['order_details']}\n" f"ID: #{order_id}")
        # küldjünk push értesítést minden regisztrált futárnak
        try:
            text = ("📣 *ÚJ RENDELÉS!* \n\n"
                    f"📍 {item['restaurant_address']}\n"
                    f"📝 {item['order_details'] or '—'}\n"
                    f"🆔 #{order_id}\n\n"
                    "Nyisd meg a futár appot és fogadd el, ha szeretnéd.")
            # sorban rakjuk be az értesítéseket; a bot worker elküldi
            notify_all_couriers_order(order_id, text)
            logger.info(f"Notification queued for order #{order_id}")
        except Exception as e:
            logger.error(f"notify couriers fail: {e}")            
            
    def notify_all_couriers_text(text: str):
        """
        A notification_queue-ba tesz be minden regisztrált futár chat_id-jére egy üzenetet.
        (A bot polling/worker majd elküldi őket.)
        """
        try:
            couriers = db.get_all_couriers()
            for c in couriers:
                uid = c.get("user_id")
                if uid:
                    notification_queue.put({"chat_id": uid, "text": text})
        except Exception as e:
            logger.error(f"notify_all_couriers_text error: {e}")

    def run(self) -> None:
        if not TELEGRAM_AVAILABLE:
            logger.warning("Telegram not available; bot not started.")
            return
        self.app.run_polling(allowed_updates=Update.ALL_TYPES)

    async def handle_callback_query(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
    
        user = update.effective_user
        if not user:
            await query.edit_message_text("Hiba: felhasználó azonosítása sikertelen")
            return
    
        data = query.data
        logger.info(f"Callback received: {data} from user {user.id}")
    
        try:
            if data.startswith("accept_"):
                # accept_ORDER_ID_MINUTES formátum
                parts = data.split("_")
                if len(parts) != 3:
                    await query.edit_message_text("Hibás parancs formátum")
                    return
                
                order_id = int(parts[1])
                eta = int(parts[2])
            
                # Rendelés elfogadása
                order = db.get_order_by_id(order_id)
                if not order or order["status"] != "pending":
                    await query.edit_message_text("Ez a rendelés már nem elérhető")
                    return
            
                partner_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or str(user.id)
                partner_username = user.username
            
                db.update_order_status(order_id, "accepted", 
                                     partner_id=user.id, 
                                     partner_name=partner_name, 
                                     partner_username=partner_username, 
                                     estimated_time=eta)
            
                # Új gombok elfogadás után
                keyboard = [
                    [InlineKeyboardButton("✅ Felvettem", callback_data=f"pickup_{order_id}")],
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
            
                await query.edit_message_text(
                    f"✅ Rendelés elfogadva!\n\n"
                    f"📍 {order['restaurant_address']}\n"
                    f"📱 {order['phone_number'] or '—'}\n"
                    f"📝 {order['order_details']}\n"
                    f"⏱️ Becsült idő: {eta} perc\n"
                    f"🆔 #{order_id}",
                    reply_markup=reply_markup
                )
            
                # Csoport értesítése
                try:
                    partner_contact = f"@{partner_username}" if partner_username else partner_name
                    group_text = (f"🚚 **FUTÁR JELENTKEZETT!**\n\n"
                                f"👤 **Futár:** {partner_name}\n"
                                f"📱 **Kontakt:** {partner_contact}\n"
                                f"⏱️ **Becsült érkezés:** {eta} perc\n"
                                f"📋 **Rendelés ID:** #{order_id}\n")
                    notification_queue.put({"chat_id": order["group_id"], "text": group_text})
                except Exception as e:
                    logger.error(f"Group notify error: {e}")
                
            elif data.startswith("pickup_"):
                order_id = int(data.split("_")[1])
            
                db.update_order_status(order_id, "picked_up", partner_id=user.id)
            
                # Új gombok felvétel után
                keyboard = [
                    [InlineKeyboardButton("✅ Kiszállítva", callback_data=f"delivered_{order_id}")],
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
            
                order = db.get_order_by_id(order_id)
                await query.edit_message_text(
                    f"📦 Rendelés felvéve!\n\n"
                    f"📍 {order['restaurant_address']}\n"
                    f"📱 {order['phone_number'] or '—'}\n"
                    f"📝 {order['order_details']}\n"
                    f"🆔 #{order_id}",
                    reply_markup=reply_markup
                )
                try:
                    partner_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or str(user.id)
                    partner_username = user.username
                    partner_contact = f"@{partner_username}" if partner_username else partner_name
        
                    group_text = (f"📦 **RENDELÉS FELVÉVE!**\n\n"
                                  f"👤 **Futár:** {partner_name}\n"
                                  f"📱 **Kontakt:** {partner_contact}\n"
                                  f"📋 **Rendelés ID:** #{order_id}\n")
                    notification_queue.put({"chat_id": order["group_id"], "text": group_text})
                except Exception as e:
                    logger.error(f"Group notify error (pickup): {e}")

            
            elif data.startswith("delivered_"):
                order_id = int(data.split("_")[1])
            
                db.update_order_status(order_id, "delivered")
            
                await query.edit_message_text(
                    f"✅ Rendelés kiszállítva!\n\n"
                    f"🆔 #{order_id}\n"
                    f"Köszönjük a munkát!"
                )
                try:
                    order = db.get_order_by_id(order_id)
                    partner_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or str(user.id)
                    partner_username = user.username
                    partner_contact = f"@{partner_username}" if partner_username else partner_name
        
                    group_text = (f"✅ **RENDELÉS KISZÁLLÍTVA!**\n\n"
                                  f"👤 **Futár:** {partner_name}\n"
                                  f"📱 **Kontakt:** {partner_contact}\n"
                                  f"📋 **Rendelés ID:** #{order_id}\n")
                    notification_queue.put({"chat_id": order["group_id"], "text": group_text})
                except Exception as e:
                    logger.error(f"Group notify error (delivered): {e}")

            
            elif data.startswith("reject_"):
                await query.edit_message_text("❌ Rendelés elutasítva")
            
        except Exception as e:
            logger.error(f"Callback error: {e}")
            await query.edit_message_text(f"Hiba történt: {str(e)}")

    async def my_orders(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        orders = db.get_orders_by_courier(user_id, status_filter="accepted")
        if not orders:
            await update.message.reply_text("📭 Nincsenek aktív rendeléseid.")
            return

        for order in orders:
            text = (f"🆔 Rendelés #{order['id']}\n"
                    f"📍 {order['restaurant_address']}\n"
                    f"📝 {order['order_details'] or '—'}\n"
                    f"Státusz: {order['status']}")
            keyboard = [
                [
                    InlineKeyboardButton("✅ Kiszállítva", callback_data=f"delivered_{order['id']}")],
                ]
            [
            await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
        ]
    
# ---------------- Flask WebApp ----------------
app = Flask(__name__); CORS(app)

def validate_telegram_data(init_data: str) -> Dict | None:
    """
    Egyszerű dekódolás a WebApp init_data-ban érkező 'user' -t tartalmazó rész alapján.
    Ha a te eredeti fájlodban más volt a dekódolás/HMAC ellenőrzés, visszaállíthatod ide.
    """
    try:
        data = {}
        for part in (init_data or "").split("&"):
            if "=" in part:
                k, v = part.split("=", 1)
                data[k] = v
        if "user" in data:
            import urllib.parse
            return json.loads(urllib.parse.unquote(data["user"]))
    except Exception as e:
        logger.error(f"validate_telegram_data error: {e}")
    return None

@app.route("/")
def index():
    try:
        orders = db.get_open_orders()
        return render_template_string(HTML_TEMPLATE, orders=orders)
    except Exception as e:
        logger.error(f"index error: {e}"); return "error", 500


HTML_TEMPLATE = r"""
<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Futár</title>
  <script src="https://telegram.org/js/telegram-web-app.js"></script>
  <style>
    body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:#fff;color:#111;margin:16px}
    .container{max-width:720px;margin:0 auto}
    .tabs {
      background: rgba(255,255,255,0.1);
      border-radius: 12px;
      padding: 4px;
    }
    .tab.active {
      background: white;
      box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    }
    @keyframes slideIn {
      from { opacity: 0; transform: translateY(20px); }
      to { opacity: 1; transform: translateY(0); }
    }
    .card { animation: slideIn 0.3s ease; }
    .card {
      background: var(--bg-card);
      border: none;
      border-radius: 16px;
      padding: 20px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(10px);
      transition: transform 0.2s ease;
    }
    .card:hover {
      transform: translateY(-2px);
    }
    .status-pending { background: linear-gradient(45deg, #fbbf24, #f59e0b); }
    .status-accepted { background: linear-gradient(45deg, #3b82f6, #1d4ed8); }
    .status-picked_up { background: linear-gradient(45deg, #8b5cf6, #7c3aed); }
    .status-delivered { background: linear-gradient(45deg, #10b981, #059669); }
    .row{display:flex;gap:8px;align-items:center;flex-wrap:wrap}
    .pill{padding:2px 8px;border-radius:999px;background:#eee;font-size:12px}
    .time-buttons{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0}
    .time-btn{border:1px solid #1a73e8;border-radius:10px;padding:10px;background:#fff;cursor:pointer;font-size:12px}
    .time-btn.selected{background:#1a73e8;color:#fff}
    .accept-btn {
      background: linear-gradient(45deg, var(--primary), var(--primary-light));
      border: none;
      border-radius: 12px;
      padding: 14px 20px;
      font-weight: 600;
      box-shadow: 0 4px 15px rgba(99, 102, 241, 0.4);
    }
    .muted{color:#666;font-size:12px}
  </style>
</head>
<body>
  <div class="container">

  <div id="admin-btn" style="display:none; margin-bottom:10px;">
    <button onclick="openAdmin()" class="accept-btn">⚙️ Admin</button>
  </div>

<script>
  function openAdmin(){
    const initData = window.Telegram?.WebApp?.initData || '';
    window.open(`${window.location.origin}/admin?init_data=${encodeURIComponent(initData)}`, '_blank');
  }

  async function checkAdmin(){
    try{
      const r = await fetch(`${window.location.origin}/api/is_admin`, {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ initData: window.Telegram?.WebApp?.initData || '' })
      });
      const j = await r.json();
      if(j.ok && j.admin){
        document.getElementById('admin-btn').style.display = 'block';
      }
    }catch(e){
      console.error('Admin check error:', e);
    }
  }
  checkAdmin();
</script>

    <h2>🍕 Futár felület</h2>

    <div class="tabs">
      <button class="tab" id="tab-av" onclick="setTab('available')">Elérhető</button>
      <button class="tab" id="tab-ac" onclick="setTab('accepted')">Elfogadott</button>
      <button class="tab" id="tab-pk" onclick="setTab('picked_up')">Felvett</button>
      <button class="tab" id="tab-dv" onclick="setTab('delivered')">Kiszállított</button>
      <button class="tab" id="tab-hist" onclick="setTab('history')">Régebbi rendelések</button>
    <div id="history-section" style="display:none;">
      <div id="weeks-list"></div>
      <div id="week-orders" style="display:none;">
        <button onclick="backToWeeks()">← Vissza a hetekhez</button>
        <div id="week-content"></div>
      </div>
    </div>

    <div class="ok" id="ok"></div>
    <div class="err" id="err"></div>
    <div id="list">Betöltés…</div>
  </div>

<script>
  const tg = window.Telegram?.WebApp; 
  if(tg) tg.expand();
  
// Regisztráljuk a futárt a szerveren, hogy a bot tudjon neki push-üzenetet küldeni
async function registerCourier(){
  try{
    const r = await fetch(`${window.location.origin}/api/register_courier`, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ initData: tg?.initData || '' })
    });
    // nem kell külön feldolgozni a választ
  }catch(e){
    console.error('registerCourier error', e);
  }
}
if(tg && tg.initData){
  registerCourier();
}
  
  const API = window.location.origin;
  let selectedETA = {}; // order_id -> 10/20/30
  let TAB = (new URLSearchParams(location.search).get('tab')) || 'available';

  function ok(m){ 
    const d=document.getElementById('ok'); 
    d.textContent=m; 
    d.style.display='block'; 
    setTimeout(()=>d.style.display='none', 3000); 
  }
  
  function err(m){ 
    const d=document.getElementById('err'); 
    d.textContent=m; 
    d.style.display='block'; 
    setTimeout(()=>d.style.display='none', 5000); 
  }
    
    const timeBtns = `
      <div class="time-buttons" style="${order.status==='pending'?'':'display:none'}">
        <button class="time-btn" data-oid="${order.id}" data-eta="10">⏱️ 10 perc</button>
        <button class="time-btn" data-oid="${order.id}" data-eta="20">⏱️ 20 perc</button>
        <button class="time-btn" data-oid="${order.id}" data-eta="30">⏱️ 30 perc</button>
      </div>
    `;
    
    let btnLabel = '🚚 Rendelés elfogadása';
    if(order.status==='accepted') btnLabel = '✅ Felvettem';
    if(order.status==='picked_up') btnLabel = '✅ Kiszállítva / Leadva';

    const showBtn = order.status !== 'delivered';
    
    return `
      <div class="card" id="card-${order.id}">
        <div class="row">
          <b>${order.group_name || order.restaurant_name}</b>
          <span class="pill">${order.status}</span>
        </div>
        <div>📍 <b>Cím:</b> ${order.restaurant_address}</div>
        ${order.phone_number ? `<div>📞 <b>Telefon:</b> ${order.phone_number}</div>` : ''}
        ${order.order_details ? `<div>📝 <b>Megjegyzés:</b> ${order.order_details}</div>` : ''}
        <div class="muted">ID: #${order.id} • ${order.created_at}</div>
        ${nav}
        ${timeBtns}
        ${showBtn ? `<button class="accept-btn" id="btn-${order.id}" onclick="doAction(${order.id}, '${order.status}')">${btnLabel}</button>` : ''}
      </div>
    `;
  }

  function wireTimeButtons(){
    document.querySelectorAll('.time-btn').forEach(b=>{
      b.addEventListener('click', ()=>{
        const oid = b.dataset.oid, eta = b.dataset.eta;
        document.querySelectorAll(`[data-oid="${oid}"]`).forEach(x=>x.classList.remove('selected'));
        b.classList.add('selected');
        selectedETA[oid] = eta;
        if(tg?.HapticFeedback) tg.HapticFeedback.impactOccurred('light');
      });
    });
  }

  async function load(){
    // tab aktív állapot
    document.getElementById('tab-av').classList.toggle('active', TAB==='available');
    document.getElementById('tab-ac').classList.toggle('active', TAB==='accepted');
    document.getElementById('tab-pk').classList.toggle('active', TAB==='picked_up');
    document.getElementById('tab-dv').classList.toggle('active', TAB==='delivered');
    

  async function doAction(orderId, status){
    const btn = document.getElementById(`btn-${orderId}`);
    if(!btn || btn.disabled) return;
    
    btn.disabled = true; 
    const old = btn.textContent; 
    btn.textContent = '⏳...';
    
    try{
      let apiUrl, payload;
      
      if(status==='pending'){
        const eta = selectedETA[orderId]; 
        if(!eta) throw new Error('Válassz időt (10/20/30 perc).');
        apiUrl = `${API}/api/accept_order`;
        payload = { order_id: orderId, estimated_time: eta, initData: tg?.initData || '' };
      } else if(status==='accepted'){
        apiUrl = `${API}/api/pickup_order`;
        payload = { order_id: orderId, initData: tg?.initData || '' };
      } else if(status==='picked_up'){
        apiUrl = `${API}/api/mark_delivered`;
        payload = { order_id: orderId, initData: tg?.initData || '' };
      } else {
        throw new Error('Ismeretlen státusz');
      }
      
      const r = await fetch(apiUrl, {
        method:'POST', 
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify(payload)
      });
      
      if (!r.ok) throw new Error(`HTTP ${r.status}: ${r.statusText}`);
      const j = await r.json(); 
      if(!j.ok) throw new Error(j.error||'Szerver hiba');
      
      // Sikeres műveletek kezelése
      if(status==='pending'){
        ok('Elfogadva.');
        btn.textContent = '✅ Felvettem';
        btn.setAttribute('onclick', `doAction(${orderId}, 'accepted')`);
        const tb = document.querySelector(`#card-${orderId} .time-buttons`); 
        if(tb) tb.style.display='none';
        const pill = document.querySelector(`#card-${orderId} .pill`); 
        if(pill) pill.textContent='accepted';
      } else if(status==='accepted'){
        ok('Felvéve.');
        btn.textContent = '✅ Kiszállítva / Leadva';
        btn.setAttribute('onclick', `doAction(${orderId}, 'picked_up')`);
        const pill = document.querySelector(`#card-${orderId} .pill`); 
        if(pill) pill.textContent='picked_up';
      } else if(status==='picked_up'){
        ok('Kiszállítva.');
        const card = document.getElementById(`card-${orderId}`);
        if(card){ 
          card.style.opacity='0.4'; 
          setTimeout(()=>card.remove(), 400); 
        }
      }
      
      btn.disabled = false;
      if(tg?.HapticFeedback) tg.HapticFeedback.notificationOccurred('success');
      
    }catch(e){
      console.error('Action error:', e);
      err(e.message || 'Hiba a művelet végrehajtásakor');
      btn.disabled = false; 
      btn.textContent = old;
      if(tg?.HapticFeedback) tg.HapticFeedback.notificationOccurred('error');
    }
  }

  // Régebbi rendelések kezelése
  async function loadHistory(){
    try{
      const r = await fetch(`${API}/api/my_orders_history`, {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ initData: tg?.initData || '' })
      });
      const j = await r.json();
      if(!j.ok) throw new Error(j.error);
      
      const weeks = j.data || [];
      const list = document.getElementById('weeks-list');
      list.innerHTML = weeks.map(w => 
        `<div class="card" onclick="loadWeekOrders('${w.week}')">
           <b>${w.week}. hét</b> - ${w.count} rendelés
         </div>`
      ).join('');
    }catch(e){
      err('Hiba a hetek betöltésénél: ' + e.message);
    }
  }

  async function loadWeekOrders(week){
    try{
      const r = await fetch(`${API}/api/my_orders_history`, {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ initData: tg?.initData || '', week: week })
      });
      const j = await r.json();
      if(!j.ok) throw new Error(j.error);
      
      const orders = j.data || [];
      document.getElementById('weeks-list').style.display = 'none';
      document.getElementById('week-orders').style.display = 'block';
      
      const content = document.getElementById('week-content');
      content.innerHTML = orders.map(render).join('');
    }catch(e){
      err('Hiba a heti rendelések betöltésénél: ' + e.message);
    }
  }

  function backToWeeks(){
    document.getElementById('weeks-list').style.display = 'block';
    document.getElementById('week-orders').style.display = 'none';
  }
  
  function setTab(t){
    TAB = t;
    
    // History section kezelése
    if(t === 'history'){
      document.getElementById('list').style.display = 'none';
      document.getElementById('history-section').style.display = 'block';
      document.getElementById('routebar').style.display = 'none';
      loadHistory();
    } else {
      document.getElementById('list').style.display = 'block';
      document.getElementById('history-section').style.display = 'none';
      load();
    }
  }

  // Kezdeti betöltés és automatikus frissítés
  load();
  setInterval(load, 30000); // 30 másodpercenként frissít
</script>
</body>
</html>
"""

ADMIN_HTML = """
<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Admin Dashboard</title>
  <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
  <style>
    :root {
      --primary: #2563eb;
      --primary-light: #3b82f6;
      --primary-dark: #1d4ed8;
      --secondary: #64748b;
      --success: #10b981;
      --warning: #f59e0b;
      --danger: #ef4444;
      --dark: #1e293b;
      --light: #f8fafc;
      --white: #ffffff;
      --border: #e2e8f0;
      --shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px 0 rgba(0, 0, 0, 0.06);
      --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
      --radius: 12px;
      --radius-lg: 16px;
    }

    * {
      margin: 0;
      padding: 0;
      box-sizing: border-box;
    }

    body {
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      min-height: 100vh;
      color: var(--dark);
      line-height: 1.6;
    }

    .container {
      max-width: 1200px;
      margin: 0 auto;
      padding: 2rem 1rem;
    }

    .header {
      background: var(--white);
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow-lg);
      padding: 2rem;
      margin-bottom: 2rem;
      display: flex;
      justify-content: space-between;
      align-items: center;
      flex-wrap: wrap;
      gap: 1rem;
    }

    .header h1 {
      font-size: 2rem;
      font-weight: 700;
      color: var(--dark);
      display: flex;
      align-items: center;
      gap: 0.75rem;
    }

    .header h1 i {
      color: var(--primary);
    }

    .export-btn {
      background: linear-gradient(135deg, var(--success) 0%, #059669 100%);
      color: var(--white);
      border: none;
      padding: 0.875rem 1.5rem;
      border-radius: var(--radius);
      font-weight: 600;
      font-size: 0.95rem;
      cursor: pointer;
      transition: all 0.2s ease;
      box-shadow: var(--shadow);
      display: flex;
      align-items: center;
      gap: 0.5rem;
      text-decoration: none;
    }

    .export-btn:hover {
      transform: translateY(-2px);
      box-shadow: var(--shadow-lg);
    }

    .stats-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
      gap: 1.5rem;
      margin-bottom: 2rem;
    }

    .stat-card {
      background: var(--white);
      border-radius: var(--radius-lg);
      padding: 1.5rem;
      box-shadow: var(--shadow);
      transition: transform 0.2s ease;
    }

    .stat-card:hover {
      transform: translateY(-2px);
    }

    .stat-header {
      display: flex;
      align-items: center;
      gap: 0.75rem;
      margin-bottom: 1rem;
    }

    .stat-icon {
      width: 48px;
      height: 48px;
      border-radius: var(--radius);
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 1.25rem;
    }

    .stat-icon.courier { background: linear-gradient(135deg, var(--primary), var(--primary-light)); color: var(--white); }
    .stat-icon.restaurant { background: linear-gradient(135deg, var(--warning), #f59e0b); color: var(--white); }
    .stat-icon.delivery { background: linear-gradient(135deg, var(--success), #059669); color: var(--white); }

    .stat-title {
      font-size: 1.125rem;
      font-weight: 600;
      color: var(--dark);
    }

    .section {
      background: var(--white);
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow);
      margin-bottom: 2rem;
      overflow: hidden;
    }

    .section-header {
      background: linear-gradient(135deg, var(--light) 0%, #e2e8f0 100%);
      padding: 1.5rem 2rem;
      border-bottom: 1px solid var(--border);
    }

    .section-title {
      font-size: 1.25rem;
      font-weight: 700;
      color: var(--dark);
      display: flex;
      align-items: center;
      gap: 0.75rem;
    }

    .section-title i {
      color: var(--primary);
    }

    .courier-block {
      border-bottom: 1px solid var(--border);
    }

    .courier-block:last-child {
      border-bottom: none;
    }

    .courier-header {
      background: var(--light);
      padding: 1rem 2rem;
      font-weight: 600;
      color: var(--dark);
      border-bottom: 1px solid var(--border);
    }

    .week-block {
      margin: 0;
    }

    .toggle-btn {
      width: 100%;
      background: var(--white);
      border: none;
      padding: 1rem 2rem;
      text-align: left;
      font-size: 0.95rem;
      font-weight: 500;
      color: var(--secondary);
      cursor: pointer;
      transition: all 0.2s ease;
      border-bottom: 1px solid var(--border);
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    .toggle-btn:hover {
      background: var(--light);
      color: var(--dark);
    }

    .toggle-btn::after {
      content: '\f107';
      font-family: 'Font Awesome 6 Free';
      font-weight: 900;
      transition: transform 0.2s ease;
    }

    .toggle-btn.active::after {
      transform: rotate(180deg);
    }

    .week-content {
      padding: 0;
      overflow: hidden;
      transition: max-height 0.3s ease;
    }

    .data-table {
      width: 100%;
      border-collapse: collapse;
    }

    .data-table th {
      background: var(--light);
      padding: 1rem;
      text-align: left;
      font-weight: 600;
      color: var(--dark);
      font-size: 0.9rem;
      text-transform: uppercase;
      letter-spacing: 0.05em;
    }

    .data-table td {
      padding: 1rem;
      border-bottom: 1px solid var(--border);
      color: var(--secondary);
    }

    .data-table tr:hover td {
      background: var(--light);
    }

    .badge {
      display: inline-flex;
      align-items: center;
      gap: 0.25rem;
      padding: 0.25rem 0.75rem;
      border-radius: 999px;
      font-size: 0.8rem;
      font-weight: 500;
    }

    .badge.success {
      background: #dcfce7;
      color: #166534;
    }

    .badge.warning {
      background: #fef3c7;
      color: #92400e;
    }

    .badge.info {
      background: #dbeafe;
      color: #1e40af;
    }

    .metric {
      font-size: 1.5rem;
      font-weight: 700;
      color: var(--primary);
    }

    .no-data {
      text-align: center;
      padding: 3rem;
      color: var(--secondary);
    }

    .no-data i {
      font-size: 3rem;
      margin-bottom: 1rem;
      opacity: 0.5;
    }

    @media (max-width: 768px) {
      .container {
        padding: 1rem 0.5rem;
      }

      .header {
        padding: 1.5rem;
        flex-direction: column;
        text-align: center;
      }

      .header h1 {
        font-size: 1.5rem;
      }

      .stats-grid {
        grid-template-columns: 1fr;
      }

      .section-header {
        padding: 1rem 1.5rem;
      }

      .courier-header, .toggle-btn {
        padding: 0.75rem 1.5rem;
      }

      .data-table {
        font-size: 0.85rem;
      }

      .data-table th, .data-table td {
        padding: 0.75rem 0.5rem;
      }
    }

    /* Animation for expanding content */
    @keyframes slideDown {
      from {
        opacity: 0;
        transform: translateY(-10px);
      }
      to {
        opacity: 1;
        transform: translateY(0);
      }
    }

    .week-content[style*="block"] {
      animation: slideDown 0.3s ease;
    }
  </style>
</head>
<body>
  <div class="container">
    <!-- Header -->
    <div class="header">
      <h1>
        <i class="fas fa-chart-line"></i>
        Admin Dashboard
      </h1>
      <form action="/admin/export_excel" method="get" style="margin: 0;">
        <input type="hidden" name="init_data" id="export-init-data">
        <button type="submit" class="export-btn">
          <i class="fas fa-download"></i>
          Excel Export
        </button>
      </form>
    </div>

    <!-- Mock Statistics Cards -->
    <div class="stats-grid">
      <div class="stat-card">
        <div class="stat-header">
          <div class="stat-icon courier">
            <i class="fas fa-motorcycle"></i>
          </div>
          <div class="stat-title">Aktív Futárok</div>
        </div>
        <div class="metric" id="active-couriers">0</div>
        <p style="color: var(--secondary); font-size: 0.9rem; margin-top: 0.5rem;">
          <i class="fas fa-arrow-up" style="color: var(--success);"></i>
          +12% az elmúlt héten
        </p>
      </div>
      
      <div class="stat-card">
        <div class="stat-header">
          <div class="stat-icon restaurant">
            <i class="fas fa-utensils"></i>
          </div>
          <div class="stat-title">Éttermek</div>
        </div>
        <div class="metric" id="restaurants">0</div>
        <p style="color: var(--secondary); font-size: 0.9rem; margin-top: 0.5rem;">
          <i class="fas fa-arrow-up" style="color: var(--success);"></i>
          +3 új regisztráció
        </p>
      </div>
      
      <div class="stat-card">
        <div class="stat-header">
          <div class="stat-icon delivery">
            <i class="fas fa-box"></i>
          </div>
          <div class="stat-title">Kiszállítások</div>
        </div>
        <div class="metric" id="deliveries">0</div>
        <p style="color: var(--secondary); font-size: 0.9rem; margin-top: 0.5rem;">
          <i class="fas fa-clock" style="color: var(--warning);"></i>
          <span id="avg-time">0</span> perc átlag
        </p>
      </div>
    </div>

    <!-- Courier Statistics Section -->
    <div class="section">
      <div class="section-header">
        <div class="section-title">
          <i class="fas fa-motorcycle"></i>
          Heti Futár Bontás
        </div>
      </div>
      <div id="courier-stats">
        {% for courier, weeks in courier_stats.items() %}
        <div class="courier-block">
          <div class="courier-header">
            <i class="fas fa-user"></i>
            {{ courier }}
          </div>
          {% for r in weeks %}
            {% set weeknum = r.week.split('-')[1] %}
            {% set year = r.week.split('-')[0] %}
            
            <div class="week-block">
              <button class="toggle-btn" onclick="toggleWeek(this)">
                <span>{{ year }}. év, {{ weeknum }}. hét</span>
                <span class="badge info">{{ r.cnt }} rendelés</span>
              </button>
              <div class="week-content" style="display: none;">
                <table class="data-table">
                  <thead>
                    <tr>
                      <th><i class="fas fa-calendar"></i> Hét</th>
                      <th><i class="fas fa-box"></i> Darab</th>
                      <th><i class="fas fa-clock"></i> Átlag Idő (perc)</th>
                    </tr>
                  </thead>
                  <tbody>
                    <tr>
                      <td>{{ r.week }}</td>
                      <td><span class="badge success">{{ r.cnt }}</span></td>
                      <td>{{ r.avg_min }}</td>
                    </tr>
                  </tbody>
                </table>
              </div>
            </div>
          {% endfor %}
        </div>
        {% endfor %}
      </div>
    </div>

    <!-- Restaurant Statistics Section -->
    <div class="section">
      <div class="section-header">
        <div class="section-title">
          <i class="fas fa-utensils"></i>
          Étterem Bontás
        </div>
      </div>
      {% if weekly_restaurant %}
      <table class="data-table">
        <thead>
          <tr>
            <th><i class="fas fa-calendar"></i> Hét</th>
            <th><i class="fas fa-store"></i> Csoport</th>
            <th><i class="fas fa-box"></i> Darab</th>
            <th><i class="fas fa-clock"></i> Átlag Idő</th>
          </tr>
        </thead>
        <tbody>
          {% for r in weekly_restaurant %}
          <tr>
            <td>{{ r.week }}</td>
            <td>{{ r.group_name }}</td>
            <td><span class="badge info">{{ r.cnt }}</span></td>
            <td>{{ r.avg_min }} perc</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
      {% else %}
      <div class="no-data">
        <i class="fas fa-chart-bar"></i>
        <p>Nincs elérhető adat</p>
      </div>
      {% endif %}
    </div>

    <!-- Detailed Deliveries Section -->
    <div class="section">
      <div class="section-header">
        <div class="section-title">
          <i class="fas fa-shipping-fast"></i>
          Részletes Kézbesítések
        </div>
      </div>
      {% if deliveries %}
      <div style="overflow-x: auto;">
        <table class="data-table">
          <thead>
            <tr>
              <th><i class="fas fa-calendar"></i> Dátum</th>
              <th><i class="fas fa-user"></i> Futár</th>
              <th><i class="fas fa-store"></i> Csoport</th>
              <th><i class="fas fa-map-marker"></i> Cím</th>
              <th><i class="fas fa-clock"></i> Idő (perc)</th>
            </tr>
          </thead>
          <tbody>
            {% for r in deliveries %}
            <tr>
              <td>{{ r.delivered_at }}</td>
              <td>{{ r.courier_name or r.delivery_partner_id }}</td>
              <td>{{ r.group_name }}</td>
              <td style="max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">
                {{ r.restaurant_address }}
              </td>
              <td>
                {% if r.min %}
                  <span class="badge {% if r.min|float < 20 %}success{% elif r.min|float < 40 %}warning{% else %}info{% endif %}">
                    {{ r.min }}
                  </span>
                {% else %}
                  <span class="badge">N/A</span>
                {% endif %}
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      {% else %}
      <div class="no-data">
        <i class="fas fa-truck"></i>
        <p>Nincs kézbesítési adat</p>
      </div>
      {% endif %}
    </div>
  </div>

  <script>
    // Set export form data
    document.getElementById("export-init-data").value = window.Telegram?.WebApp?.initData || '';

    // Toggle week details
    function toggleWeek(btn) {
      const content = btn.nextElementSibling;
      const isVisible = content.style.display === "block";
      
      if (isVisible) {
        content.style.display = "none";
        btn.classList.remove('active');
      } else {
        content.style.display = "block";
        btn.classList.add('active');
      }
    }

    // Mock data for statistics (replace with actual data from your backend)
    function updateStats() {
      // These would be populated from your actual data
      const courierStats = {{ courier_stats|tojson }};
      const deliveryStats = {{ deliveries|tojson }};
      const restaurantStats = {{ weekly_restaurant|tojson }};

      // Count unique couriers
      const uniqueCouriers = Object.keys(courierStats || {}).length;
      document.getElementById('active-couriers').textContent = uniqueCouriers;

      // Count unique restaurants
      const uniqueRestaurants = new Set((restaurantStats || []).map(r => r.group_name)).size;
      document.getElementById('restaurants').textContent = uniqueRestaurants;

      // Total deliveries and average time
      const totalDeliveries = (deliveryStats || []).length;
      document.getElementById('deliveries').textContent = totalDeliveries;

      if (deliveryStats && deliveryStats.length > 0) {
        const avgTime = deliveryStats
          .filter(d => d.min)
          .reduce((sum, d) => sum + parseFloat(d.min), 0) / 
          deliveryStats.filter(d => d.min).length;
        document.getElementById('avg-time').textContent = Math.round(avgTime);
      }
    }

    // Delete order function for admin
    async function deleteOrder(orderId) {
      if (!confirm(`Biztosan törölni szeretnéd a #${orderId} rendelést?`)) {
        return;
      }
      
      try {
        const response = await fetch('/api/delete_order', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({
            initData: window.Telegram?.WebApp?.initData || '',
            order_id: orderId
          })
        });
        
        const data = await response.json();
        if (data.ok) {
          alert('Rendelés sikeresen törölve!');
          location.reload(); // Refresh the page to show updated data
        } else {
          alert('Hiba a törlés során: ' + (data.error || 'Ismeretlen hiba'));
        }
      } catch (e) {
        alert('Hálózati hiba a törlés során: ' + e.message);
      }
    }

    // Initialize stats
    updateStats();

    // Add smooth scrolling
    document.querySelectorAll('a[href^="#"]').forEach(anchor => {
      anchor.addEventListener('click', function (e) {
        e.preventDefault();
        document.querySelector(this.getAttribute('href')).scrollIntoView({
          behavior: 'smooth'
        });
      });
    });
  </script>
</body>
</html>
"""

# --- other endpoints preserved (accept/pickup/mark_delivered etc) ---
@app.route("/api/orders")
def api_orders():
    try: return jsonify(db.get_open_orders())
    except Exception as e: logger.error(f"api_orders error: {e}"); return jsonify([])

@app.route("/api/accept_order", methods=["POST"])
def api_accept_order():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        eta = int(data.get("estimated_time", 20))
        user = validate_telegram_data(data.get("initData", ""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        order = db.get_order_by_id(order_id)
        if not order or order["status"] != "pending": return jsonify({"ok": False, "error": "not_available"}), 400
        partner_name = ((user.get("first_name","") + " " + user.get("last_name","")).strip()) or str(user.get("id"))
        partner_username = user.get("username")
        db.update_order_status(order_id, "accepted", partner_id=user.get("id"), partner_name=partner_name, partner_username=partner_username, estimated_time=eta)
        try:
            partner_contact = f"@{partner_username}" if partner_username else partner_name
            text = ("🚚 **FUTÁR JELENTKEZETT!**\n\n" f"👤 **Futár:** {partner_name}\n" f"📱 **Kontakt:** {partner_contact}\n" f"⏱️ **Becsült érkezés:** {eta} perc\n" f"📋 **Rendelés ID:** #{order_id}\n")
            notification_queue.put({"chat_id": order["group_id"], "text": text})
        except Exception as e: logger.error(f"group notify fail (accept): {e}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_accept_order error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/pickup_order", methods=["POST"])
def api_pickup_order():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        order = db.get_order_by_id(order_id)
        if not order or order["status"] != "accepted" or order.get("delivery_partner_id") not in (None, user.get("id")):
            pass
        db.update_order_status(order_id, "picked_up", partner_id=user.get("id"))
        try:
            partner_name = ((user.get("first_name","") + " " + user.get("last_name","")).strip()) or str(user.get("id"))
            partner_username = user.get("username"); partner_contact = f"@{partner_username}" if partner_username else partner_name
            text = ("📦 **RENDELÉS FELVÉVE!**\n\n" f"👤 **Futár:** {partner_name}\n" f"📱 **Kontakt:** {partner_contact}\n" f"📋 **Rendelés ID:** #{order_id}\n")
            notification_queue.put({"chat_id": order["group_id"], "text": text})
        except Exception as e: logger.error(f"group notify fail (pickup): {e}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_pickup_order error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/mark_delivered", methods=["POST"])
def api_mark_delivered():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        order = db.get_order_by_id(order_id)
        if not order or order["status"] != "picked_up": return jsonify({"ok": False, "error": "not_pickup"}), 400
        db.update_order_status(order_id, "delivered")
        try:
            partner_name = ((user.get("first_name","") + " " + user.get("last_name","")).strip()) or str(user.get("id"))
            partner_username = user.get("username"); partner_contact = f"@{partner_username}" if partner_username else partner_name
            text = ("✅ **RENDELÉS KISZÁLLÍTVA!**\n\n" f"👤 **Futár:** {partner_name}\n" f"📱 **Kontakt:** {partner_contact}\n" f"📋 **Rendelés ID:** #{order_id}\n")
            notification_queue.put({"chat_id": order["group_id"], "text": text})
        except Exception as e: logger.error(f"group notify fail (delivered): {e}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_mark_delivered error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/orders_by_status", methods=["GET"])
def api_orders_by_status():
    try:
        status = (request.args.get("status") or "").strip()
        courier_id = request.args.get("courier_id", type=int)
        if not status: return jsonify([])
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        if status == "pending":
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status FROM orders WHERE status='pending' AND DATE(created_at) = DATE('now', 'localtime') ORDER BY created_at DESC")
            rows = [dict(r) for r in cur.fetchall()]
        elif status in ("accepted","picked_up","delivered"):
            if not courier_id: conn.close(); return jsonify({"ok": False, "error": "missing_courier"}), 400
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, estimated_time FROM orders WHERE status=? AND delivery_partner_id=? ORDER BY created_at DESC", (status, courier_id))
            rows = [dict(r) for r in cur.fetchall()]
        else:
            rows = []
        conn.close(); return jsonify(rows)
    except Exception as e:
        logger.error(f"api_orders_by_status error: {e}"); return jsonify([]), 500

@app.route("/api/my_orders", methods=["POST"])
def api_my_orders():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error":"unauthorized"}), 401
        status = data.get("status","").strip()
        if status not in ("accepted","picked_up","delivered"): return jsonify({"ok": True, "orders": []})
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        if status == "delivered":
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, estimated_time FROM orders WHERE status=? AND delivery_partner_id=? AND DATE(created_at) = DATE('now', 'localtime') ORDER BY created_at DESC", (status, user["id"]))
        else:
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, estimated_time FROM orders WHERE status=? AND delivery_partner_id=? ORDER BY created_at DESC", (status, user["id"]))
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify({"ok": True, "orders": rows})
    except Exception as e:
        logger.error(f"api_my_orders error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500
    
@app.route("/admin")
def admin_page():
    init_data = request.args.get('init_data', ''); user = validate_telegram_data(init_data)
    if not user or user.get("id") not in ADMIN_USER_IDS: return "🚫 Hozzáférés megtagadva", 403
    try:
        import datetime
        current_week = datetime.date.today().strftime("%Y-%W")
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("""
            SELECT strftime('%Y-%W', delivered_at) AS week, delivery_partner_id, COALESCE(delivery_partner_name, '') AS courier_name, COUNT(*) AS cnt, ROUND(AVG((julianday(delivered_at) - julianday(accepted_at)) * 24 * 60), 1) AS avg_min FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL GROUP BY delivery_partner_id, week ORDER BY week DESC, cnt DESC
        """)
        weekly_courier = [dict(r) for r in cur.fetchall()]
        from collections import defaultdict
        courier_stats = defaultdict(list)
        for r in weekly_courier:
            courier_stats[r["courier_name"] or r["delivery_partner_id"]].append(r)
        cur.execute("""
            SELECT strftime('%Y-%W', delivered_at) AS week, group_name, COUNT(*) AS cnt, ROUND(AVG((julianday(delivered_at) - julianday(accepted_at)) * 24 * 60), 1) AS avg_min FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL GROUP BY group_name, week ORDER BY week DESC, cnt DESC
        """)
        weekly_restaurant = [dict(r) for r in cur.fetchall()]
        cur.execute("""
            SELECT delivered_at, delivery_partner_id, COALESCE(delivery_partner_name, '') AS courier_name, group_name, restaurant_address, ROUND((julianday(delivered_at) - julianday(accepted_at)) * 24 * 60, 1) AS min FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL ORDER BY delivered_at DESC LIMIT 500
        """)
        deliveries = [dict(r) for r in cur.fetchall()]; conn.close()
        return render_template_string(ADMIN_HTML,
                                      courier_stats=courier_stats,
                                      weekly_restaurant=weekly_restaurant,
                                      deliveries=deliveries,
                                      current_week=current_week)
    except Exception as e:
        logger.error(f"admin_page error: {e}"); return "admin error", 500

@app.route("/api/is_admin", methods=["POST"])
def api_is_admin():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData", ""))
        if not user: return jsonify({"ok": False, "admin": False}), 401
        return jsonify({"ok": True, "admin": user.get("id") in ADMIN_USER_IDS})
    except Exception as e:
        logger.error(f"api_is_admin error: {e}"); return jsonify({"ok": False, "admin": False}), 500

@app.route("/admin/export_excel")
def admin_export_excel():

    conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()

    # ugyanaz a 3 lekérdezés, mint admin_page-ben
    cur.execute("""SELECT strftime('%Y-%W', delivered_at) AS week, delivery_partner_id, 
                   COALESCE(delivery_partner_name, '') AS courier_name, COUNT(*) AS cnt,
                   ROUND(AVG((julianday(delivered_at)-julianday(accepted_at))*24*60),1) AS avg_min
                   FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL
                   GROUP BY delivery_partner_id, week ORDER BY week DESC, cnt DESC""")
    weekly_courier = [dict(r) for r in cur.fetchall()]

    cur.execute("""SELECT strftime('%Y-%W', delivered_at) AS week, group_name, COUNT(*) AS cnt,
                   ROUND(AVG((julianday(delivered_at)-julianday(accepted_at))*24*60),1) AS avg_min
                   FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL
                   GROUP BY group_name, week ORDER BY week DESC, cnt DESC""")
    weekly_restaurant = [dict(r) for r in cur.fetchall()]

    cur.execute("""SELECT delivered_at, delivery_partner_id, COALESCE(delivery_partner_name, '') AS courier_name,
                   group_name, restaurant_address,
                   ROUND((julianday(delivered_at)-julianday(accepted_at))*24*60,1) AS min
                   FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL
                   ORDER BY delivered_at DESC LIMIT 500""")
    deliveries = [dict(r) for r in cur.fetchall()]
    conn.close()

    # Excel létrehozás
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Heti futár"
    ws1.append(["Hét", "Futár", "Darab", "Átlag idő (perc)"])
    for r in weekly_courier:
        ws1.append([r["week"], r["courier_name"] or r["delivery_partner_id"], r["cnt"], r["avg_min"]])

    ws2 = wb.create_sheet("Heti étterem")
    ws2.append(["Hét", "Csoport", "Darab", "Átlag idő (perc)"])
    for r in weekly_restaurant:
        ws2.append([r["week"], r["group_name"], r["cnt"], r["avg_min"]])

    # Éttermek szétválogatása
    deliveries_by_restaurant = {}
    for r in deliveries:
        gname = r["group_name"] or "Ismeretlen"
        deliveries_by_restaurant.setdefault(gname, []).append(r)

    # Minden étteremhez külön munkalap
    for gname, rows in deliveries_by_restaurant.items():
       # Excel sheet neve max 31 karakter lehet → levágjuk, ha hosszabb
        safe_name = gname[:31]
        ws = wb.create_sheet(safe_name)
        ws.append(["Dátum", "Futár", "Csoport", "Cím", "Idő (perc)"])
        for r in rows:
            ws.append([r["delivered_at"], r["courier_name"] or r["delivery_partner_id"],
                       r["group_name"], r["restaurant_address"], r["min"]])
    # mentés memóriába
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name="statisztika.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/my_orders_history", methods=["POST"])
def api_my_orders_history():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error":"unauthorized"}), 401
        
        week = data.get("week", "").strip()  # formátum: "2025-01"
        
        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        if week:
            cur.execute("""
                SELECT id, restaurant_name, restaurant_address, phone_number, order_details, 
                       group_id, group_name, created_at, status, estimated_time, delivered_at
                FROM orders 
                WHERE delivery_partner_id=? AND strftime('%Y-%W', created_at) = ?
                ORDER BY created_at DESC
            """, (user["id"], week))
        else:
            # Összes hét listázása
            cur.execute("""
                SELECT DISTINCT strftime('%Y-%W', created_at) as week, 
                       COUNT(*) as count
                FROM orders 
                WHERE delivery_partner_id=?
                GROUP BY strftime('%Y-%W', created_at)
                ORDER BY week DESC
            """, (user["id"],))
        
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"ok": True, "data": rows})
    except Exception as e:
        logger.error(f"api_my_orders_history error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/register_courier", methods=["POST"])
def api_register_courier():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData",""))
        if not user:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
        db.register_courier(user)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_register_courier error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/get_address", methods=["POST"])
def api_get_address():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        user = validate_telegram_data(data.get("initData",""))
        if not user: 
            return jsonify({"ok": False, "error": "unauthorized"}), 401
        
        order = db.get_order_by_id(order_id)
        if not order:
            return jsonify({"ok": False, "error": "order_not_found"}), 404
            
        return jsonify({
            "ok": True, 
            "address": order.get("restaurant_address","")
        })
    except Exception as e:
        logger.error(f"api_get_address error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------- Server start ----------------
def run_flask():
    from werkzeug.serving import WSGIRequestHandler
    WSGIRequestHandler.protocol_version = "HTTP/1.1"
    app.run(host="0.0.0.0", port=5000, debug=False)

if __name__ == "__main__":
    # start flask in background thread and start bot polling (if available)
    threading.Thread(target=run_flask, daemon=True).start()
    if TELEGRAM_AVAILABLE:
        RestaurantBot().run()
    else:
        logger.info("Flask started in this environment; Telegram bot not available.")
