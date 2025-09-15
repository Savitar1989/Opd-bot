
# file: opdtest_final.py
# NOTE: Replace BOT_TOKEN and WEBAPP_URL with your real values before running.
import asyncio
import os
import logging
import sqlite3
import json
import threading
import re
import requests
import urllib.parse
import math
import itertools
import time
from queue import Queue, Empty
from typing import Dict, List, Optional, Tuple

from flask import Flask, render_template_string, request, jsonify
from flask_cors import CORS

from flask import send_file
import openpyxl
from io import BytesIO

from collections import defaultdict


# Telegram imports are optional if you run bot; keep them to preserve original behavior
try:
    from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
    import urllib.parse
    from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
    TELEGRAM_AVAILABLE = True
except Exception:
    TELEGRAM_AVAILABLE = False

# =============== CONFIG ===============
BOT_TOKEN = "7741178469:AAH9pvClqBOa31Yenq_0Y9dxtrug-ZMmDk4"
WEBAPP_URL = "https://94377687755d.ngrok-free.app"
DB_NAME = "restaurant_orders.db"
ADMIN_USER_IDS = [7553912440]  # adjust as needed

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

notification_queue: "Queue[Dict]" = Queue()

# ---------------- Utilities: Address parsing / Geocode ----------------
def parse_hungarian_address(address: str) -> str:
    """
    Kezdeti egyszerű normalizálás: eltávolít felesleges szóközöket és egyes rövidítéseket.
    Konkrét, hosszabb eredeti parser megtartva a projektedben, itt minimalizálva.
    """
    if not address:
        return ""
    addr = address.strip()
    # egyszerű normalizálás: több szóköz -> egy
    addr = re.sub(r'\s+', ' ', addr)
    return addr

def geocode_address(address: str) -> Optional[Tuple[float, float]]:
    """
    Geokódolás Nominatim szolgáltatással. Visszatér (lat, lon) vagy None.
    """
    try:
        time.sleep(0.4)
        parsed = parse_hungarian_address(address)
        url = "https://nominatim.openstreetmap.org/search"
        params = {'q': parsed, 'format': 'json', 'limit': 1, 'countrycodes': 'hu', 'addressdetails': 1}
        headers = {'User-Agent': 'OPDRouteBot/1.0'}
        r = requests.get(url, params=params, headers=headers, timeout=8)
        if r.status_code == 200:
            data = r.json()
            if data and len(data) > 0:
                return (float(data[0]['lat']), float(data[0]['lon']))
    except Exception as e:
        logger.error(f"geocode error for '{address}': {e}")
    return None

# ---------------- Distance & TSP helpers ----------------
def haversine_distance(a: Tuple[float, float], b: Tuple[float, float]) -> float:
    lat1, lon1 = a; lat2, lon2 = b
    R = 6371.0
    dlat = math.radians(lat2 - lat1); dlon = math.radians(lon2 - lon1)
    sa = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(sa))
    return R * c

def calculate_total_distance(route: List[Tuple[str, float, float]]) -> float:
    if not route or len(route) < 2: return 0.0
    total = 0.0
    for i in range(len(route)-1):
        total += haversine_distance((route[i][1], route[i][2]), (route[i+1][1], route[i+1][2]))
    return total

def rotate_route_to_centroid_start(route: List[Tuple[str, float, float]]) -> List[Tuple[str, float, float]]:
    """
    Forgatjuk az útvonalat, hogy a kezdő pont a centroid-hoz legközelebb legyen.
    Ezt csak akkor használjuk, ha nincs expliciten megadott start (futár helye).
    """
    if not route:
        return route
    lat_sum = sum(r[1] for r in route)
    lon_sum = sum(r[2] for r in route)
    centroid = (lat_sum / len(route), lon_sum / len(route))
    min_idx = 0; min_d = float('inf')
    for i, r in enumerate(route):
        d = haversine_distance((r[1], r[2]), centroid)
        if d < min_d:
            min_d = d; min_idx = i
    return route[min_idx:] + route[:min_idx]

def tsp_2opt(coords_with_addr: List[Tuple[str, float, float]]) -> List[Tuple[str, float, float]]:
    if len(coords_with_addr) <= 2:
        return coords_with_addr
    
    n = len(coords_with_addr)
    has_fixed_start = coords_with_addr[0][0] == "CURRENT_LOCATION"
    
    # Ha van rögzített start, azt megtartjuk
    if has_fixed_start:
        fixed_start = coords_with_addr[0]
        points = coords_with_addr[1:]
    else:
        fixed_start = None
        points = coords_with_addr[:]
    
    if not points:
        return coords_with_addr
    
    # Távolság mátrix létrehozása
    def distance_matrix(pts):
        matrix = {}
        for i, p1 in enumerate(pts):
            for j, p2 in enumerate(pts):
                if i != j:
                    matrix[(i, j)] = haversine_distance((p1[1], p1[2]), (p2[1], p2[2]))
                else:
                    matrix[(i, j)] = 0
        return matrix
    
    # Ha van rögzített start, azt is beletesszük a mátrixba
    all_points = [fixed_start] + points if fixed_start else points
    distances = {}
    for i, p1 in enumerate(all_points):
        for j, p2 in enumerate(all_points):
            if i != j:
                distances[(i, j)] = haversine_distance((p1[1], p1[2]), (p2[1], p2[2]))
            else:
                distances[(i, j)] = 0
    
    # Nearest neighbor algoritmus javított változata
    if fixed_start:
        route = [0]  # Start a rögzített ponttal
        unvisited = list(range(1, len(all_points)))
    else:
        # Legjobb start pont keresése: válasszuk a centroidhoz legközelebbit
        lat_center = sum(p[1] for p in all_points) / len(all_points)
        lon_center = sum(p[2] for p in all_points) / len(all_points)
        best_start = min(range(len(all_points)), 
                        key=lambda i: haversine_distance((all_points[i][1], all_points[i][2]), 
                                                        (lat_center, lon_center)))
        route = [best_start]
        unvisited = [i for i in range(len(all_points)) if i != best_start]
    
    # Nearest neighbor építés
    while unvisited:
        current = route[-1]
        next_city = min(unvisited, key=lambda city: distances[(current, city)])
        route.append(next_city)
        unvisited.remove(next_city)
    
    # 2-opt javítás
    def two_opt_swap(route, i, k):
        new_route = route[:]
        new_route[i:k+1] = route[i:k+1][::-1]
        return new_route
    
    def route_distance(route):
        total = 0
        for i in range(len(route)):
            j = (i + 1) % len(route)
            total += distances[(route[i], route[j])]
        return total
    
    # 2-opt optimalizálás
    improved = True
    max_iterations = 1000
    iteration = 0
    
    while improved and iteration < max_iterations:
        improved = False
        iteration += 1
        
        for i in range(len(route)):
            for k in range(i + 2, len(route)):
                # Ha van rögzített start, ne mozgassuk
                if fixed_start and (i == 0 or k == len(route) - 1):
                    continue
                    
                new_route = two_opt_swap(route, i, k)
                if route_distance(new_route) < route_distance(route):
                    route = new_route
                    improved = True
                    break
            if improved:
                break
    
    # Vissza alakítás koordinátákra
    return [all_points[i] for i in route]

def optimize_route(addresses: List[str], start_coord: Optional[Tuple[str,float,float]] = None) -> List[Tuple[str,float,float]]:
    """
    Geokódol minden címet (ha lehetséges), majd optimalizálja a sorrendet.
    start_coord: optional ("CURRENT_LOCATION", lat, lon) amely mindig az első elem lesz.
    Visszaadott lista: [(address, lat, lon), ...] - első elem a start, ha volt.
    """
    if not addresses:
        return []
    if len(addresses) > 12:
        addresses = addresses[:12]  # korlátozás
    coords_with_addr = []
    for a in addresses:
        c = geocode_address(a)
        if c:
            coords_with_addr.append((a, c[0], c[1]))
        else:
            logger.warning(f"Could not geocode: {a}")
    if not coords_with_addr:
        return []
    # Insert start coordinate if provided
    if start_coord:
        coords_with_addr.insert(0, start_coord)
    # Small n: brute-force permutations (including start if present)
    if len(coords_with_addr) <= 5:
        best = list(coords_with_addr)
        min_d = float('inf')
        for perm in itertools.permutations(coords_with_addr[1:] if start_coord else coords_with_addr):
            candidate = ([coords_with_addr[0]] + list(perm)) if start_coord else list(perm)
            d = calculate_total_distance(candidate)
            if d < min_d:
                min_d = d; best = candidate
        if not start_coord:
            best = rotate_route_to_centroid_start(best)
        return best
    # 2-opt optimalizálás
    optimized = tsp_2opt(coords_with_addr)
    if not start_coord:
        optimized = rotate_route_to_centroid_start(optimized)
    return optimized

# ---------------- Map URL builders ----------------
def coords_to_google_maps_url(coords_with_addr: List[Tuple[str, float, float]]) -> str:
    """
    Koordináták alapján Google Maps URL generálása.
    - Több cím esetén köztes pontokat ad hozzá.
    - Azonnal indítja a navigációt mobilon (&dir_action=navigate).
    """
    if not coords_with_addr:
        return ""

    if len(coords_with_addr) == 1:
        lat, lon = coords_with_addr[0][1], coords_with_addr[0][2]
        return f"https://www.google.com/maps/search/?api=1&query={lat},{lon}&dir_action=navigate"

    # Utolsó pont = célállomás
    dest_lat, dest_lon = coords_with_addr[-1][1], coords_with_addr[-1][2]
    destination = f"{dest_lat},{dest_lon}"

    # Köztes pontok
    waypoints = [f"{lat},{lon}" for addr, lat, lon in coords_with_addr[:-1]]
    waypoints_str = "|".join(waypoints)

    return (
        f"https://www.google.com/maps/dir/?api=1"
        f"&destination={destination}"
        f"&waypoints={waypoints_str}"
        f"&travelmode=driving"
        f"&dir_action=navigate"
    )

def coords_to_apple_maps_url(coords_with_addr: List[Tuple[str, float, float]]) -> str:
    if not coords_with_addr:
        return ""
    # Apple accepts multiple daddr params; origin isn't necessary for Apple dir links but include for safety
    parts = []
    for i, (_, lat, lon) in enumerate(coords_with_addr):
        if i == 0:
            parts.append(f"saddr={lat},{lon}")
        else:
            parts.append(f"daddr={lat},{lon}")
    return "https://maps.apple.com/?" + "&".join(parts) + "&dirflg=d"

def coords_to_waze_url(coords_with_addr: List[Tuple[str, float, float]]) -> str:
    if not coords_with_addr:
        return ""
    lat, lon = coords_with_addr[0][1], coords_with_addr[0][2]
    return f"https://waze.com/ul?ll={lat},{lon}&navigate=yes"

# ---------------- Database Manager ----------------
class DatabaseManager:
    def __init__(self) -> None:
        self.init_db()
        self.db_path = DB_NAME
    def init_db(self) -> None:
        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()
        # Összes tábla létrehozása egy kapcsolaton belül
        cur.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                restaurant_name TEXT NOT NULL,
                restaurant_address TEXT NOT NULL,
                phone_number TEXT,
                order_details TEXT NOT NULL,
                group_id INTEGER NOT NULL,
                group_name TEXT,
                message_id INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT (datetime('now', 'localtime')),
                status TEXT DEFAULT 'pending',
                delivery_partner_id INTEGER,
                delivery_partner_name TEXT,
                delivery_partner_username TEXT,
                estimated_time INTEGER,
                accepted_at TIMESTAMP,
                picked_up_at TIMESTAMP,
                delivered_at TIMESTAMP
            )
        """)
    
        cur.execute("""CREATE TABLE IF NOT EXISTS groups(id INTEGER PRIMARY KEY, name TEXT NOT NULL)""")
    
        cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_group ON orders(group_name)")
    
        # Futárok tábla létrehozása ugyanabban a kapcsolatban
        cur.execute("""CREATE TABLE IF NOT EXISTS couriers (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                first_name TEXT,
                last_name TEXT,
                last_seen TIMESTAMP DEFAULT (datetime('now', 'localtime'))
            )
        """)
    
        # Minden változtatás mentése és kapcsolat bezárása
        conn.commit()
        conn.close()

    def register_group(self, group_id: int, group_name: str) -> None:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("INSERT OR IGNORE INTO groups(id, name) VALUES (?,?)", (group_id, group_name))
        conn.commit(); conn.close()

    def save_order(self, item: Dict) -> int:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("""INSERT INTO orders (restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, message_id) VALUES (?,?,?,?,?,?,?)""",
                    (item.get("restaurant_name",""), item.get("restaurant_address",""), item.get("phone_number",""), item.get("order_details",""), item.get("group_id"), item.get("group_name"), item.get("message_id")))
        oid = cur.lastrowid; conn.commit(); conn.close(); return oid

    def get_open_orders(self) -> List[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, delivery_partner_id, estimated_time FROM orders WHERE status IN ('pending','accepted','picked_up') ORDER BY created_at DESC")
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return rows

    def get_order_by_id(self, order_id: int) -> Optional[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT * FROM orders WHERE id = ?", (order_id,)); row = cur.fetchone(); conn.close(); return dict(row) if row else None

    def update_order_status(self, order_id: int, status: str, partner_id: int | None = None, partner_name: str | None = None, partner_username: str | None = None, estimated_time: int | None = None) -> None:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("""
            UPDATE orders SET status = ?, delivery_partner_id = COALESCE(?, delivery_partner_id), delivery_partner_name = COALESCE(?, delivery_partner_name), delivery_partner_username = COALESCE(?, delivery_partner_username), estimated_time = COALESCE(?, estimated_time), accepted_at = CASE WHEN ?='accepted' THEN datetime('now', 'localtime') ELSE accepted_at END, picked_up_at = CASE WHEN ?='picked_up' THEN datetime ('now', 'localtime') ELSE picked_up_at END, delivered_at = CASE WHEN ?='delivered' THEN datetime('now', 'localtime') ELSE delivered_at END WHERE id = ?
        """, (status, partner_id, partner_name, partner_username, estimated_time, status, status, status, order_id))
        conn.commit(); conn.close()

    def get_partner_addresses(self, partner_id: int, status: str) -> List[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT id, restaurant_address, group_name FROM orders WHERE delivery_partner_id = ? AND status = ? ORDER BY created_at", (partner_id, status))
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return rows

    def get_partner_order_count(self, partner_id: int, status: str = None) -> int:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        if status: cur.execute("SELECT COUNT(*) FROM orders WHERE delivery_partner_id = ? AND status = ?", (partner_id, status))
        else: cur.execute("SELECT COUNT(*) FROM orders WHERE delivery_partner_id = ?", (partner_id,))
        c = cur.fetchone()[0]; conn.close(); return c

    def register_courier(self, user: Dict) -> None:
        conn = sqlite3.connect(DB_NAME); cur = conn.cursor()
        cur.execute("""
            INSERT OR REPLACE INTO couriers(user_id, username, first_name, last_name, last_seen) 
            VALUES (?,?,?,?,datetime('now','localtime'))
        """, (user.get("id"), user.get("username"), user.get("first_name"), user.get("last_name")))
        conn.commit(); conn.close()

    def get_all_couriers(self) -> List[Dict]:
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("SELECT user_id, username, first_name, last_name FROM couriers")
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return rows

db = DatabaseManager()

def notify_all_couriers_order(order_id: int, text: str):
    """
    Interaktív értesítés küldése inline keyboard-dal
    """
    try:
        from telegram import InlineKeyboardButton, InlineKeyboardMarkup
        
        # Gombok létrehozása
        keyboard = [
            [
                InlineKeyboardButton("⏱️ 10 perc", callback_data=f"accept_{order_id}_10"),
                InlineKeyboardButton("⏱️ 20 perc", callback_data=f"accept_{order_id}_20"),
                InlineKeyboardButton("⏱️ 30 perc", callback_data=f"accept_{order_id}_30")
            ],
            [InlineKeyboardButton("❌ Elutasítás", callback_data=f"reject_{order_id}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        couriers = db.get_all_couriers()
        logger.info(f"Found {len(couriers)} registered couriers")
        for c in couriers:
            uid = c.get("user_id")
            if uid:
                notification_queue.put({
                    "chat_id": uid, 
                    "text": text,
                    "reply_markup": reply_markup
                })
                logger.info(f"Queued interactive notification for courier {uid}")
    except Exception as e:
        logger.error(f"notify_all_couriers_order error: {e}")

def get_orders_by_courier(self, courier_id, status_filter="accepted"):
    with sqlite3.connect(DB_NAME) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM orders WHERE partner_id = ? AND status = ?",
            (courier_id, status_filter),
        )
        return [dict(row) for row in cur.fetchall()]
        
# ---------------- Telegram Bot (kept intact) ----------------
class RestaurantBot:
    def __init__(self) -> None:
        if not TELEGRAM_AVAILABLE:
            logger.warning("python-telegram-bot not available; bot handlers won't be active in this environment.")
            return
        self.app = Application.builder().token(BOT_TOKEN).build()
        self._setup_handlers()

    def _setup_handlers(self) -> None:
        app = self.app
        app.add_handler(CommandHandler("start", self.start_cmd))
        app.add_handler(CommandHandler("help", self.help_cmd))
        app.add_handler(CommandHandler("register", self.register_group))
        app.add_handler(MessageHandler(filters.TEXT & filters.ChatType.GROUPS, self.handle_group_message))
        app.add_handler(CommandHandler("myorders", self.my_orders))
        app.add_handler(CommandHandler("route_all", self.route_all))
        app.add_handler(CommandHandler("route", self.route_single))
        app.add_handler(CallbackQueryHandler(self.handle_callback_query))
    
        if app.job_queue:
            app.job_queue.run_repeating(self.process_notifications, interval=3)

    async def process_notifications(self, context: ContextTypes.DEFAULT_TYPE):
        processed_count = 0; max_per_batch = 5
        while processed_count < max_per_batch:
            try: 
                item = notification_queue.get_nowait()
                processed_count += 1
                logger.info(f"Processing notification for chat_id: {item.get('chat_id')}")
            except Empty: 
                if processed_count == 0:
                    logger.debug("No notifications in queue")
                break
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    # Inline keyboard támogatás hozzáadása
                    kwargs = {
                        "chat_id": item["chat_id"],
                        "text": item.get("text", ""),
                        "parse_mode": "Markdown"
                    }
                    if "reply_markup" in item:
                        kwargs["reply_markup"] = item["reply_markup"]
                
                    await context.bot.send_message(**kwargs)
                    logger.info(f"Successfully sent notification to {item['chat_id']}")
                    break
                except Exception as e:
                    logger.error(f"Failed to send notification to {item['chat_id']} (attempt {attempt+1}): {e}")
                    if attempt < max_retries - 1: await asyncio.sleep(1)

    def send_notification(self, chat_id: int, text: str):
        try:
            if not text or not chat_id: return
            notification_queue.put({"chat_id": chat_id, "text": text})
        except Exception as e:
            logger.error(f"Error queueing notification: {e}")

    async def start_cmd(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user
        if update.effective_chat.type == "private":
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("🚚 Elérhető rendelések", web_app=WebAppInfo(url=f"{WEBAPP_URL}"))]])
            await update.message.reply_text(f"Üdv, {user.first_name}!\nNyisd meg a futár felületet:", reply_markup=kb)
        else:
            await update.message.reply_text("Használd a /register parancsot a csoport regisztrálásához.\nRendelés formátum:\nCím: ...\nTelefonszám: ...\nMegjegyzés: ...")

    async def help_cmd(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text("Rendelés formátum (csoportban):\n```\nCím: Budapest, Példa utca 1.\nTelefonszám: +36301234567\nMegjegyzés: kp / kártya / megjegyzés\n```", parse_mode="Markdown")

    async def register_group(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_chat.type not in ("group","supergroup"):
            await update.message.reply_text("Ezt a parancsot csoportban használd."); return
        gid = update.effective_chat.id; gname = update.effective_chat.title or "Ismeretlen csoport"
        db.register_group(gid, gname); await update.message.reply_text(f"✅ A '{gname}' csoport regisztrálva.")

    def parse_order_message(self, text: str) -> Dict | None:
        lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
        info = {}
        def after_colon(s: str) -> str:
            return s.split(":",1)[1].strip() if ":" in s else ""
        for ln in lines:
            low = ln.lower()
            if low.startswith("cím:") or low.startswith("cim:"): info["address"] = after_colon(ln)
            elif low.startswith("telefonszám:") or low.startswith("telefonszam:") or low.startswith("telefon:"): info["phone"] = after_colon(ln)
            elif low.startswith("megjegyzés:") or low.startswith("megjegyzes:"): info["details"] = after_colon(ln)
        if info.get("address"): info.setdefault("phone",""); info.setdefault("details",""); return info
        return None

    async def handle_group_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_chat.type not in ("group","supergroup"): return
        parsed = self.parse_order_message(update.message.text or ""); 
        if not parsed: return
        gid = update.effective_chat.id; gname = update.effective_chat.title or "Ismeretlen"
        item = {"restaurant_name": gname, "restaurant_address": parsed["address"], "phone_number": parsed.get("phone",""), "order_details": parsed.get("details",""), "group_id": gid, "group_name": gname, "message_id": update.message.message_id}
        order_id = db.save_order(item)
        await update.message.reply_text("✅ Rendelés rögzítve.\n\n" f"📍 Cím: {item['restaurant_address']}\n" f"📞 Telefon: {item['phone_number'] or '—'}\n" f"📝 Megjegyzés: {item['order_details']}\n" f"ID: #{order_id}")
        # küldjünk push értesítést minden regisztrált futárnak
        try:
            text = ("📣 *ÚJ RENDELÉS!* \n\n"
                    f"📍 {item['restaurant_address']}\n"
                    f"📝 {item['order_details'] or '—'}\n"
                    f"🆔 #{order_id}\n\n"
                    "Nyisd meg a futár appot és fogadd el, ha szeretnéd.")
            # sorban rakjuk be az értesítéseket; a bot worker elküldi
            notify_all_couriers_order(order_id, text)
            logger.info(f"Notification queued for order #{order_id}")
        except Exception as e:
            logger.error(f"notify couriers fail: {e}")            
            
    def notify_all_couriers_text(text: str):
        """
        A notification_queue-ba tesz be minden regisztrált futár chat_id-jére egy üzenetet.
        (A bot polling/worker majd elküldi őket.)
        """
        try:
            couriers = db.get_all_couriers()
            for c in couriers:
                uid = c.get("user_id")
                if uid:
                    notification_queue.put({"chat_id": uid, "text": text})
        except Exception as e:
            logger.error(f"notify_all_couriers_text error: {e}")

    def run(self) -> None:
        if not TELEGRAM_AVAILABLE:
            logger.warning("Telegram not available; bot not started.")
            return
        self.app.run_polling(allowed_updates=Update.ALL_TYPES)

    async def handle_callback_query(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
    
        user = update.effective_user
        if not user:
            await query.edit_message_text("Hiba: felhasználó azonosítása sikertelen")
            return
    
        data = query.data
        logger.info(f"Callback received: {data} from user {user.id}")
    
        try:
            if data.startswith("accept_"):
                # accept_ORDER_ID_MINUTES formátum
                parts = data.split("_")
                if len(parts) != 3:
                    await query.edit_message_text("Hibás parancs formátum")
                    return
                
                order_id = int(parts[1])
                eta = int(parts[2])
            
                # Rendelés elfogadása
                order = db.get_order_by_id(order_id)
                if not order or order["status"] != "pending":
                    await query.edit_message_text("Ez a rendelés már nem elérhető")
                    return
            
                partner_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or str(user.id)
                partner_username = user.username
            
                db.update_order_status(order_id, "accepted", 
                                     partner_id=user.id, 
                                     partner_name=partner_name, 
                                     partner_username=partner_username, 
                                     estimated_time=eta)
            
                # Új gombok elfogadás után
                keyboard = [
                    [InlineKeyboardButton("✅ Felvettem", callback_data=f"pickup_{order_id}")],
                    [InlineKeyboardButton("🗺️ Navigáció", callback_data=f"navigate_{order_id}")]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
            
                await query.edit_message_text(
                    f"✅ Rendelés elfogadva!\n\n"
                    f"📍 {order['restaurant_address']}\n"
                    f"📱 {order['phone_number'] or '—'}\n"
                    f"📝 {order['order_details']}\n"
                    f"⏱️ Becsült idő: {eta} perc\n"
                    f"🆔 #{order_id}",
                    reply_markup=reply_markup
                )
            
                # Csoport értesítése
                try:
                    partner_contact = f"@{partner_username}" if partner_username else partner_name
                    group_text = (f"🚚 **FUTÁR JELENTKEZETT!**\n\n"
                                f"👤 **Futár:** {partner_name}\n"
                                f"📱 **Kontakt:** {partner_contact}\n"
                                f"⏱️ **Becsült érkezés:** {eta} perc\n"
                                f"📋 **Rendelés ID:** #{order_id}\n")
                    notification_queue.put({"chat_id": order["group_id"], "text": group_text})
                except Exception as e:
                    logger.error(f"Group notify error: {e}")
                
            elif data.startswith("pickup_"):
                order_id = int(data.split("_")[1])
            
                db.update_order_status(order_id, "picked_up", partner_id=user.id)
            
                # Új gombok felvétel után
                keyboard = [
                    [InlineKeyboardButton("✅ Kiszállítva", callback_data=f"delivered_{order_id}")],
                    [InlineKeyboardButton("🗺️ Navigáció", callback_data=f"navigate_{order_id}")]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
            
                order = db.get_order_by_id(order_id)
                await query.edit_message_text(
                    f"📦 Rendelés felvéve!\n\n"
                    f"📍 {order['restaurant_address']}\n"
                    f"📱 {order['phone_number'] or '—'}\n"
                    f"📝 {order['order_details']}\n"
                    f"🆔 #{order_id}",
                    reply_markup=reply_markup
                )
            
            elif data.startswith("delivered_"):
                order_id = int(data.split("_")[1])
            
                db.update_order_status(order_id, "delivered")
            
                await query.edit_message_text(
                    f"✅ Rendelés kiszállítva!\n\n"
                    f"🆔 #{order_id}\n"
                    f"Köszönjük a munkát!"
                )

            elif data == "route_all":
                await self.route_all(update, context)
            
            elif data.startswith("navigate_"):
                order_id = int(data.split("_")[1])
                order = db.get_order_by_id(order_id)
            
                if order:
                    coord = geocode_address(order["restaurant_address"])
                    if coord:
                        lat, lon = coord
                        google_url = f"https://www.google.com/maps/search/?api=1&query={lat},{lon}&dir_action=navigate"
                    
                        # Navigációs gombok
                        nav_keyboard = [
                            [InlineKeyboardButton("🗺️ Google Maps", url=google_url)],
                            [InlineKeyboardButton("🍎 Apple Maps", url=f"https://maps.apple.com/?daddr={lat},{lon}&dirflg=d")],
                            [InlineKeyboardButton("🚗 Waze", url=f"https://waze.com/ul?ll={lat},{lon}&navigate=yes")],
                            [InlineKeyboardButton("⬅️ Vissza", callback_data=f"back_{order_id}")]
                        ]
                        nav_markup = InlineKeyboardMarkup(nav_keyboard)
                    
                        await query.edit_message_text(
                            f"🗺️ Navigáció indítása:\n\n"
                            f"📍 {order['restaurant_address']}\n"
                            f"🆔 #{order_id}",
                            reply_markup=nav_markup
                        )
                    else:
                        await query.answer("Hiba a geokódolásban", show_alert=True)
            
            elif data.startswith("reject_"):
                await query.edit_message_text("❌ Rendelés elutasítva")
            
        except Exception as e:
            logger.error(f"Callback error: {e}")
            await query.edit_message_text(f"Hiba történt: {str(e)}")

    async def my_orders(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        orders = db.get_orders_by_courier(user_id, status_filter="accepted")
        if not orders:
            await update.message.reply_text("📭 Nincsenek aktív rendeléseid.")
            return

        for order in orders:
            text = (f"🆔 Rendelés #{order['id']}\n"
                    f"📍 {order['restaurant_address']}\n"
                    f"📝 {order['order_details'] or '—'}\n"
                    f"Státusz: {order['status']}")
            keyboard = [
                [
                    InlineKeyboardButton("✅ Kiszállítva", callback_data=f"delivered_{order['id']}"),
                    InlineKeyboardButton("📍 Navigáció", url=f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(order['restaurant_address'])}")
                ]
            ]
            await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))

        # Összes rendeléshez útvonal gomb
        all_keyboard = [
            [InlineKeyboardButton("🗺 Útvonal az összeshez", callback_data="route_all")
        ]
    ]
        await update.message.reply_text("Összes rendeléshez útvonal:", reply_markup=InlineKeyboardMarkup(all_keyboard))


    async def route_all(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        orders = db.get_orders_by_courier(user_id, status_filter="accepted")
        if not orders:
            await update.message.reply_text("📭 Nincsenek aktív rendeléseid.")
            return

        addresses = [o["restaurant_address"] for o in orders]
        # Url-kódolás miatt
        encoded = [urllib.parse.quote(addr) for addr in addresses]
        if len(encoded) == 1:
            maps_url = f"https://www.google.com/maps/search/?api=1&query={encoded[0]}"
        else:
            waypoints = "|".join(encoded[:-1])
            destination = encoded[-1]
            maps_url = f"https://www.google.com/maps/dir/?api=1&destination={destination}&waypoints={waypoints}&travelmode=driving"
            
        await update.message.reply_text(f"🗺 Útvonal minden rendeléshez:\n{maps_url}")


    async def route_single(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not context.args or len(context.args) < 1:
            await update.message.reply_text("Használat: /route <rendeles_id>")
            return
        try:
            order_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Az ID-nak számnak kell lennie.")
            return

        order = db.get_order_by_id(order_id)
        if not order:
            await update.message.reply_text("❌ Nincs ilyen rendelés.")
            return

        addr = order["restaurant_address"]
        maps_url = f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(addr)}"
        await update.message.reply_text(f"🗺 Útvonal a rendeléshez:\n{maps_url}")

# ---------------- Flask WebApp ----------------
app = Flask(__name__); CORS(app)

def validate_telegram_data(init_data: str) -> Dict | None:
    """
    Egyszerű dekódolás a WebApp init_data-ban érkező 'user' -t tartalmazó rész alapján.
    Ha a te eredeti fájlodban más volt a dekódolás/HMAC ellenőrzés, visszaállíthatod ide.
    """
    try:
        data = {}
        for part in (init_data or "").split("&"):
            if "=" in part:
                k, v = part.split("=", 1)
                data[k] = v
        if "user" in data:
            import urllib.parse
            return json.loads(urllib.parse.unquote(data["user"]))
    except Exception as e:
        logger.error(f"validate_telegram_data error: {e}")
    return None

# NOTE: Replace the HTML_TEMPLATE with your original HTML content if needed.
HTML_TEMPLATE = r"""<!-- PLACE YOUR ORIGINAL HTML TEMPLATE HERE -->"""

@app.route("/")
def index():
    try:
        orders = db.get_open_orders()
        return render_template_string(HTML_TEMPLATE, orders=orders)
    except Exception as e:
        logger.error(f"index error: {e}"); return "error", 500

@app.route("/api/get_coordinates", methods=["POST"])
def api_get_coordinates():
    """
    Returns plain numeric coordinates for a given order_id.
    Body: { order_id: <int>, initData: <tg.initData> }
    """
    try:
        data = request.json or {}
        order_id = int(data.get("order_id", 0))
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        if not order_id: return jsonify({"ok": False, "error": "missing_order_id"}), 400
        order = db.get_order_by_id(order_id)
        if not order: return jsonify({"ok": False, "error": "order_not_found"}), 404
        coord = geocode_address(order.get("restaurant_address",""))
        if not coord: return jsonify({"ok": False, "error": "geocode_failed"}), 500
        lat, lon = coord
        return jsonify({"ok": True, "lat": lat, "lon": lon})
    except Exception as e:
        logger.error(f"api_get_coordinates error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/optimize_route", methods=["POST"])
def api_optimize_route():
    """
    Optimizes route for orders assigned to the courier (status='picked_up').
    Accepts optional 'current_lat' and 'current_lon' in the request body — if present, they are used as the start point.
    Returns coordinate-only list and prebuilt Google Maps URL.
    """
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData", ""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        rows = db.get_partner_addresses(partner_id=user["id"], status="picked_up")
        addresses = [r["restaurant_address"] for r in rows if r.get("restaurant_address")]
        if not addresses: return jsonify({"ok": False, "error": "no_addresses"}), 400
        # parse provided current position (optional) - prefer explicit start
        start_coord = None
        try:
            if data.get("current_lat") is not None and data.get("current_lon") is not None:
                start_coord = ("CURRENT_LOCATION", float(data.get("current_lat")), float(data.get("current_lon")))
        except Exception:
            start_coord = None
        optimized = optimize_route(addresses, start_coord=start_coord)
        # ensure optimized contains coords_only in string form for client
        coords_list = [f"{lat},{lon}" for (_addr, lat, lon) in optimized]
        coords_objects = [{"address": _addr, "lat": lat, "lon": lon} for (_addr, lat, lon) in optimized]
        google_url = coords_to_google_maps_url(optimized)
        apple_url = coords_to_apple_maps_url(optimized)
        waze_url = coords_to_waze_url(optimized)
        return jsonify({"ok": True, "addresses": coords_list, "coords": coords_objects, "google_url": google_url, "apple_url": apple_url, "waze_url": waze_url, "count": len(coords_list)})
    except Exception as e:
        logger.error(f"api_optimize_route error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500


HTML_TEMPLATE = r"""
<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Futár</title>
  <script src="https://telegram.org/js/telegram-web-app.js"></script>
  <style>
    body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:#fff;color:#111;margin:16px}
    .container{max-width:720px;margin:0 auto}
    .tabs {
      background: rgba(255,255,255,0.1);
      border-radius: 12px;
      padding: 4px;
    }
    .tab.active {
      background: white;
      box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    }
    @keyframes slideIn {
      from { opacity: 0; transform: translateY(20px); }
      to { opacity: 1; transform: translateY(0); }
    }
    .card { animation: slideIn 0.3s ease; }
    .card {
      background: var(--bg-card);
      border: none;
      border-radius: 16px;
      padding: 20px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(10px);
      transition: transform 0.2s ease;
    }
    .card:hover {
      transform: translateY(-2px);
    }
    .status-pending { background: linear-gradient(45deg, #fbbf24, #f59e0b); }
    .status-accepted { background: linear-gradient(45deg, #3b82f6, #1d4ed8); }
    .status-picked_up { background: linear-gradient(45deg, #8b5cf6, #7c3aed); }
    .status-delivered { background: linear-gradient(45deg, #10b981, #059669); }
    .row{display:flex;gap:8px;align-items:center;flex-wrap:wrap}
    .pill{padding:2px 8px;border-radius:999px;background:#eee;font-size:12px}
    .time-buttons{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin:10px 0}
    .time-btn{border:1px solid #1a73e8;border-radius:10px;padding:10px;background:#fff;cursor:pointer;font-size:12px}
    .time-btn.selected{background:#1a73e8;color:#fff}
    .accept-btn {
      background: linear-gradient(45deg, var(--primary), var(--primary-light));
      border: none;
      border-radius: 12px;
      padding: 14px 20px;
      font-weight: 600;
      box-shadow: 0 4px 15px rgba(99, 102, 241, 0.4);
    }
    .muted{color:#666;font-size:12px}
    
    /* Navigációs gombok stílusai */
    .nav-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin:8px 0}
    .nav{display:block;text-decoration:none;border:1px solid #1a73e8;border-radius:8px;padding:8px;background:#fff;text-align:center;font-size:11px;color:#1a73e8}
    .nav.apple{border-color:#000;color:#000;background:#f5f5f7}
    .nav.waze{border-color:#33ccff;color:#33ccff;background:#f0f8ff}
    .nav:hover{opacity:0.8}
    
    .ok{display:none;background:#d4edda;color:#155724;border-radius:8px;padding:10px;margin:8px 0}
    .err{display:none;background:#f8d7da;color:#721c24;border-radius:8px;padding:10px;margin:8px 0}
    
    /* Útvonal optimalizáló gombok */
    .routebar{display:none;gap:6px;margin:8px 0;flex-wrap:wrap}
    .routebtn{border:0;border-radius:10px;padding:10px 12px;background:#1a73e8;color:#fff;cursor:pointer;font-size:12px}
    .routebtn.apple{background:#000}
    .routebtn.waze{background:#33ccff}
    .routebtn:hover{opacity:0.9}
  </style>
</head>
<body>
  <div class="container">

  <div id="admin-btn" style="display:none; margin-bottom:10px;">
    <button onclick="openAdmin()" class="accept-btn">⚙️ Admin</button>
  </div>

<script>
  function openAdmin(){
    const initData = window.Telegram?.WebApp?.initData || '';
    window.open(`${window.location.origin}/admin?init_data=${encodeURIComponent(initData)}`, '_blank');
  }

  async function checkAdmin(){
    try{
      const r = await fetch(`${window.location.origin}/api/is_admin`, {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ initData: window.Telegram?.WebApp?.initData || '' })
      });
      const j = await r.json();
      if(j.ok && j.admin){
        document.getElementById('admin-btn').style.display = 'block';
      }
    }catch(e){
      console.error('Admin check error:', e);
    }
  }
  checkAdmin();
</script>

    <h2>🍕 Futár felület</h2>

    <div class="tabs">
      <button class="tab" id="tab-av" onclick="setTab('available')">Elérhető</button>
      <button class="tab" id="tab-ac" onclick="setTab('accepted')">Elfogadott</button>
      <button class="tab" id="tab-pk" onclick="setTab('picked_up')">Felvett</button>
      <button class="tab" id="tab-dv" onclick="setTab('delivered')">Kiszállított</button>
      <button class="tab" id="tab-hist" onclick="setTab('history')">Régebbi rendelések</button>
    <div id="history-section" style="display:none;">
      <div id="weeks-list"></div>
      <div id="week-orders" style="display:none;">
        <button onclick="backToWeeks()">← Vissza a hetekhez</button>
        <div id="week-content"></div>
      </div>
    </div>

    <!-- Navigációs gombok - csak Felvett menüben -->
    <div class="routebar" id="routebar" style="display:none;">
      <button class="routebtn" onclick="openOptimizedRoute('google')">🗺️ Google Maps - Optimalizált útvonal</button>
      <button class="routebtn apple" onclick="openOptimizedRoute('apple')">🍎 Apple Maps - Optimalizált útvonal</button>
      <button class="routebtn waze" onclick="openOptimizedRoute('waze')">🚗 Waze - Optimalizált útvonal</button>
    </div>

    <div class="ok" id="ok"></div>
    <div class="err" id="err"></div>
    <div id="list">Betöltés…</div>
  </div>

<script>
  const tg = window.Telegram?.WebApp; 
  if(tg) tg.expand();
  
// Regisztráljuk a futárt a szerveren, hogy a bot tudjon neki push-üzenetet küldeni
async function registerCourier(){
  try{
    const r = await fetch(`${window.location.origin}/api/register_courier`, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ initData: tg?.initData || '' })
    });
    // nem kell külön feldolgozni a választ
  }catch(e){
    console.error('registerCourier error', e);
  }
}
if(tg && tg.initData){
  registerCourier();
}
  
  const API = window.location.origin;
  let selectedETA = {}; // order_id -> 10/20/30
  let TAB = (new URLSearchParams(location.search).get('tab')) || 'available';

  function ok(m){ 
    const d=document.getElementById('ok'); 
    d.textContent=m; 
    d.style.display='block'; 
    setTimeout(()=>d.style.display='none', 3000); 
  }
  
  function err(m){ 
    const d=document.getElementById('err'); 
    d.textContent=m; 
    d.style.display='block'; 
    setTimeout(()=>d.style.display='none', 5000); 
  }

  // Navigációs függvények
  // HELPERS: cím tisztítása / dekódolása, hibabiztos
    function normalizeAddress(addr){
      if(!addr && addr !== 0) return '';
      try {
    // ha %-kódolt részeket találunk, próbáljuk dekódolni (pl. 'Danko%20Pista' -> 'Danko Pista')
        if (/%[0-9A-Fa-f]{2}/.test(addr)) {
            addr = decodeURIComponent(addr);
        }
      } catch(e) {
    // ha a dekódolás hibát dob (hibás %xx), hagyjuk az eredetit
      }
  // pluszokból szóköz, többszörös whitespace normalizálás, trim
      addr = String(addr).replace(/\+/g, ' ').replace(/\s+/g, ' ').trim();
      return addr;
    }

 async function googleMapsLink(orderId){
  try {
    const r = await fetch(`${API}/api/get_coordinates`, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ order_id: orderId, initData: tg?.initData || '' })
    });
    const j = await r.json();
    if(j.ok) {
      return `https://www.google.com/maps/search/?api=1&query=${j.lat},${j.lon}&dir_action=navigate`;
    }
  } catch(e) {
    console.error('Koordináta lekérési hiba:', e);
  }
  return '#'; // fallback
}

async function appleMapsLink(orderId){
  try {
    const r = await fetch(`${API}/api/get_coordinates`, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ order_id: orderId, initData: tg?.initData || '' })
    });
    const j = await r.json();
    if(j.ok) {
      return `https://maps.apple.com/?daddr=${j.lat},${j.lon}&dirflg=d`;
    }
  } catch(e) {
    console.error('Koordináta lekérési hiba:', e);
  }
  return '#'; // fallback
}

async function wazeLink(orderId){
  try {
    const r = await fetch(`${API}/api/get_coordinates`, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ order_id: orderId, initData: tg?.initData || '' })
    });
    const j = await r.json();
    if(j.ok) {
      return `https://waze.com/ul?ll=${j.lat},${j.lon}&navigate=yes`;
    }
  } catch(e) {
    console.error('Koordináta lekérési hiba:', e);
  }
  return '#'; // fallback
}

  function render(order){
    // Navigációs gombok - csak Felvett menüben
    const nav = (TAB === 'picked_up') ? `
      <div class="nav-grid">
        <a class="nav" href="#" onclick="openGoogleMaps(${order.id})" target="_blank">🗺️ Google</a>
        <a class="nav apple" href="#" onclick="openAppleMaps(${order.id})" target="_blank">🍎 Apple</a>
        <a class="nav waze" href="#" onclick="openWaze(${order.id})" target="_blank">🚗 Waze</a>
      </div>
    ` : '';
    
    const timeBtns = `
      <div class="time-buttons" style="${order.status==='pending'?'':'display:none'}">
        <button class="time-btn" data-oid="${order.id}" data-eta="10">⏱️ 10 perc</button>
        <button class="time-btn" data-oid="${order.id}" data-eta="20">⏱️ 20 perc</button>
        <button class="time-btn" data-oid="${order.id}" data-eta="30">⏱️ 30 perc</button>
      </div>
    `;
    
    let btnLabel = '🚚 Rendelés elfogadása';
    if(order.status==='accepted') btnLabel = '✅ Felvettem';
    if(order.status==='picked_up') btnLabel = '✅ Kiszállítva / Leadva';

    const showBtn = order.status !== 'delivered';
    
    return `
      <div class="card" id="card-${order.id}">
        <div class="row">
          <b>${order.group_name || order.restaurant_name}</b>
          <span class="pill">${order.status}</span>
        </div>
        <div>📍 <b>Cím:</b> ${order.restaurant_address}</div>
        ${order.phone_number ? `<div>📞 <b>Telefon:</b> ${order.phone_number}</div>` : ''}
        ${order.order_details ? `<div>📝 <b>Megjegyzés:</b> ${order.order_details}</div>` : ''}
        <div class="muted">ID: #${order.id} • ${order.created_at}</div>
        ${nav}
        ${timeBtns}
        ${showBtn ? `<button class="accept-btn" id="btn-${order.id}" onclick="doAction(${order.id}, '${order.status}')">${btnLabel}</button>` : ''}
      </div>
    `;
  }

  async function openGoogleMaps(orderId) {
      const link = await googleMapsLink(orderId);
      if(link !== '#') window.open(link, '_blank');
  }

  async function openAppleMaps(orderId) {
      const link = await appleMapsLink(orderId);
      if(link !== '#') window.open(link, '_blank');
  }

  async function openWaze(orderId) {
      const link = await wazeLink(orderId);
      if(link !== '#') window.open(link, '_blank');
  }

  function wireTimeButtons(){
    document.querySelectorAll('.time-btn').forEach(b=>{
      b.addEventListener('click', ()=>{
        const oid = b.dataset.oid, eta = b.dataset.eta;
        document.querySelectorAll(`[data-oid="${oid}"]`).forEach(x=>x.classList.remove('selected'));
        b.classList.add('selected');
        selectedETA[oid] = eta;
        if(tg?.HapticFeedback) tg.HapticFeedback.impactOccurred('light');
      });
    });
  }

  async function load(){
    // tab aktív állapot
    document.getElementById('tab-av').classList.toggle('active', TAB==='available');
    document.getElementById('tab-ac').classList.toggle('active', TAB==='accepted');
    document.getElementById('tab-pk').classList.toggle('active', TAB==='picked_up');
    document.getElementById('tab-dv').classList.toggle('active', TAB==='delivered');
    
    // Navigációs gombok megjelenítése csak Felvett menüben
    document.getElementById('routebar').style.display = (TAB==='picked_up') ? 'flex' : 'none';

    const list = document.getElementById('list');
    list.innerHTML = 'Betöltés…';

    let data = [];
    try{
      if(TAB === 'available'){
        const r = await fetch(`${API}/api/orders_by_status?status=pending`);
        if (!r.ok) throw new Error(`HTTP ${r.status}: ${r.statusText}`);
        data = await r.json();
      }else{
        const r = await fetch(`${API}/api/my_orders`, {
          method:'POST', 
          headers:{'Content-Type':'application/json'},
          body: JSON.stringify({ initData: tg?.initData || '', status: TAB })
        });
        if (!r.ok) throw new Error(`HTTP ${r.status}: ${r.statusText}`);
        const j = await r.json();
        if(!j.ok) throw new Error(j.error||'Hálózati hiba');
        data = j.orders || [];
      }
    }catch(e){
      console.error('Load error:', e);
      err(e.message||'Hiba a betöltésnél');
      data = [];
    }

    if(!data.length){ 
      list.innerHTML = '<div class="muted">Nincs rendelés.</div>'; 
      return; 
    }
    
    list.innerHTML = data.map(render).join('');
    wireTimeButtons();
  }

  async function doAction(orderId, status){
    const btn = document.getElementById(`btn-${orderId}`);
    if(!btn || btn.disabled) return;
    
    btn.disabled = true; 
    const old = btn.textContent; 
    btn.textContent = '⏳...';
    
    try{
      let apiUrl, payload;
      
      if(status==='pending'){
        const eta = selectedETA[orderId]; 
        if(!eta) throw new Error('Válassz időt (10/20/30 perc).');
        apiUrl = `${API}/api/accept_order`;
        payload = { order_id: orderId, estimated_time: eta, initData: tg?.initData || '' };
      } else if(status==='accepted'){
        apiUrl = `${API}/api/pickup_order`;
        payload = { order_id: orderId, initData: tg?.initData || '' };
      } else if(status==='picked_up'){
        apiUrl = `${API}/api/mark_delivered`;
        payload = { order_id: orderId, initData: tg?.initData || '' };
      } else {
        throw new Error('Ismeretlen státusz');
      }
      
      const r = await fetch(apiUrl, {
        method:'POST', 
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify(payload)
      });
      
      if (!r.ok) throw new Error(`HTTP ${r.status}: ${r.statusText}`);
      const j = await r.json(); 
      if(!j.ok) throw new Error(j.error||'Szerver hiba');
      
      // Sikeres műveletek kezelése
      if(status==='pending'){
        ok('Elfogadva.');
        btn.textContent = '✅ Felvettem';
        btn.setAttribute('onclick', `doAction(${orderId}, 'accepted')`);
        const tb = document.querySelector(`#card-${orderId} .time-buttons`); 
        if(tb) tb.style.display='none';
        const pill = document.querySelector(`#card-${orderId} .pill`); 
        if(pill) pill.textContent='accepted';
      } else if(status==='accepted'){
        ok('Felvéve.');
        btn.textContent = '✅ Kiszállítva / Leadva';
        btn.setAttribute('onclick', `doAction(${orderId}, 'picked_up')`);
        const pill = document.querySelector(`#card-${orderId} .pill`); 
        if(pill) pill.textContent='picked_up';
      } else if(status==='picked_up'){
        ok('Kiszállítva.');
        const card = document.getElementById(`card-${orderId}`);
        if(card){ 
          card.style.opacity='0.4'; 
          setTimeout(()=>card.remove(), 400); 
        }
      }
      
      btn.disabled = false;
      if(tg?.HapticFeedback) tg.HapticFeedback.notificationOccurred('success');
      
    }catch(e){
      console.error('Action error:', e);
      err(e.message || 'Hiba a művelet végrehajtásakor');
      btn.disabled = false; 
      btn.textContent = old;
      if(tg?.HapticFeedback) tg.HapticFeedback.notificationOccurred('error');
    }
  }

  // Útvonal optimalizáló függvény
    async function openOptimizedRoute(mapType = 'google'){
      try{
        const r = await fetch(`${API}/api/optimize_route`, {
          method:'POST',
          headers:{'Content-Type':'application/json'},
          body: JSON.stringify({ initData: tg?.initData || '' })
        });

        if (!r.ok) throw new Error(`HTTP ${r.status}: ${r.statusText}`);
        const j = await r.json();
        if(!j.ok) throw new Error(j.error||'Hálózati hiba');

        // Use coords array from server (objects with {address, lat, lon})
        const coords = j.coords || [];
        if(coords.length === 0){
          err('Nincs felvett rendelés az útvonaltervezéshez');
          return;
        }
      
      // Navigációs URL generálása
      let url;
      if(mapType === 'apple'){
      // Apple Maps: daddr=lat,lon repeated
      const daddr_params = coords.map(c => `daddr=${encodeURIComponent(c.address)}`).join('&');
      url = `https://maps.apple.com/?${daddr_params}&dirflg=d`;

     } else if(mapType === 'waze'){
       // Waze: use first point (Waze web supports ll=lat,lon)
      const c = coords[0];
      url = `https://waze.com/ul?q=${encodeURIComponent(c.address)}&navigate=yes`;
     } else {
        // Google Maps - optimalizált útvonal
        if(coords.length === 1){
        url = `https://www.google.com/maps/search/?api=1&query=${coords[0].lat},${coords[0].lon}&dir_action=navigate`;
      } else {
        const destination = `${coords[coords.length-1].lat},${coords[coords.length-1].lon}`;
        const waypoints = coords.slice(0,-1).map(c => `${c.lat},${c.lon}`).join('|');
        url = `https://www.google.com/maps/dir/?api=1&destination=${destination}&waypoints=${waypoints}&travelmode=driving&dir_action=navigate`;
        }
    }

      // Link megnyitása
      if (tg?.openLink) {
        
        tg.openLink(url);
      } else {
        window.open(url, '_blank');
      }
      
      const mapNames = {
        'google': 'Google Maps',
        'apple': 'Apple Maps', 
        'waze': 'Waze'
      };
      
      ok(`${mapNames[mapType]} megnyitva ${addresses.length} optimalizált címmel`);
      
    }catch(e){
      console.error('Route error:', e);
      err(e.message || 'Hiba az útvonaltervezésnél');
    }
  }

  // Régebbi rendelések kezelése
  async function loadHistory(){
    try{
      const r = await fetch(`${API}/api/my_orders_history`, {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ initData: tg?.initData || '' })
      });
      const j = await r.json();
      if(!j.ok) throw new Error(j.error);
      
      const weeks = j.data || [];
      const list = document.getElementById('weeks-list');
      list.innerHTML = weeks.map(w => 
        `<div class="card" onclick="loadWeekOrders('${w.week}')">
           <b>${w.week}. hét</b> - ${w.count} rendelés
         </div>`
      ).join('');
    }catch(e){
      err('Hiba a hetek betöltésénél: ' + e.message);
    }
  }

  async function loadWeekOrders(week){
    try{
      const r = await fetch(`${API}/api/my_orders_history`, {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ initData: tg?.initData || '', week: week })
      });
      const j = await r.json();
      if(!j.ok) throw new Error(j.error);
      
      const orders = j.data || [];
      document.getElementById('weeks-list').style.display = 'none';
      document.getElementById('week-orders').style.display = 'block';
      
      const content = document.getElementById('week-content');
      content.innerHTML = orders.map(render).join('');
    }catch(e){
      err('Hiba a heti rendelések betöltésénél: ' + e.message);
    }
  }

  function backToWeeks(){
    document.getElementById('weeks-list').style.display = 'block';
    document.getElementById('week-orders').style.display = 'none';
  }
  
  function setTab(t){
    TAB = t;
    
    // History section kezelése
    if(t === 'history'){
      document.getElementById('list').style.display = 'none';
      document.getElementById('history-section').style.display = 'block';
      document.getElementById('routebar').style.display = 'none';
      loadHistory();
    } else {
      document.getElementById('list').style.display = 'block';
      document.getElementById('history-section').style.display = 'none';
      load();
    }
  }

  // Kezdeti betöltés és automatikus frissítés
  load();
  setInterval(load, 30000); // 30 másodpercenként frissít
</script>
</body>
</html>
"""

ADMIN_HTML = """
<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Admin Dashboard</title>
  <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
  <style>
    :root {
      --primary: #2563eb;
      --primary-light: #3b82f6;
      --primary-dark: #1d4ed8;
      --secondary: #64748b;
      --success: #10b981;
      --warning: #f59e0b;
      --danger: #ef4444;
      --dark: #1e293b;
      --light: #f8fafc;
      --white: #ffffff;
      --border: #e2e8f0;
      --shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.1), 0 1px 2px 0 rgba(0, 0, 0, 0.06);
      --shadow-lg: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
      --radius: 12px;
      --radius-lg: 16px;
    }

    * {
      margin: 0;
      padding: 0;
      box-sizing: border-box;
    }

    body {
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      min-height: 100vh;
      color: var(--dark);
      line-height: 1.6;
    }

    .container {
      max-width: 1200px;
      margin: 0 auto;
      padding: 2rem 1rem;
    }

    .header {
      background: var(--white);
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow-lg);
      padding: 2rem;
      margin-bottom: 2rem;
      display: flex;
      justify-content: space-between;
      align-items: center;
      flex-wrap: wrap;
      gap: 1rem;
    }

    .header h1 {
      font-size: 2rem;
      font-weight: 700;
      color: var(--dark);
      display: flex;
      align-items: center;
      gap: 0.75rem;
    }

    .header h1 i {
      color: var(--primary);
    }

    .export-btn {
      background: linear-gradient(135deg, var(--success) 0%, #059669 100%);
      color: var(--white);
      border: none;
      padding: 0.875rem 1.5rem;
      border-radius: var(--radius);
      font-weight: 600;
      font-size: 0.95rem;
      cursor: pointer;
      transition: all 0.2s ease;
      box-shadow: var(--shadow);
      display: flex;
      align-items: center;
      gap: 0.5rem;
      text-decoration: none;
    }

    .export-btn:hover {
      transform: translateY(-2px);
      box-shadow: var(--shadow-lg);
    }

    .stats-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
      gap: 1.5rem;
      margin-bottom: 2rem;
    }

    .stat-card {
      background: var(--white);
      border-radius: var(--radius-lg);
      padding: 1.5rem;
      box-shadow: var(--shadow);
      transition: transform 0.2s ease;
    }

    .stat-card:hover {
      transform: translateY(-2px);
    }

    .stat-header {
      display: flex;
      align-items: center;
      gap: 0.75rem;
      margin-bottom: 1rem;
    }

    .stat-icon {
      width: 48px;
      height: 48px;
      border-radius: var(--radius);
      display: flex;
      align-items: center;
      justify-content: center;
      font-size: 1.25rem;
    }

    .stat-icon.courier { background: linear-gradient(135deg, var(--primary), var(--primary-light)); color: var(--white); }
    .stat-icon.restaurant { background: linear-gradient(135deg, var(--warning), #f59e0b); color: var(--white); }
    .stat-icon.delivery { background: linear-gradient(135deg, var(--success), #059669); color: var(--white); }

    .stat-title {
      font-size: 1.125rem;
      font-weight: 600;
      color: var(--dark);
    }

    .section {
      background: var(--white);
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow);
      margin-bottom: 2rem;
      overflow: hidden;
    }

    .section-header {
      background: linear-gradient(135deg, var(--light) 0%, #e2e8f0 100%);
      padding: 1.5rem 2rem;
      border-bottom: 1px solid var(--border);
    }

    .section-title {
      font-size: 1.25rem;
      font-weight: 700;
      color: var(--dark);
      display: flex;
      align-items: center;
      gap: 0.75rem;
    }

    .section-title i {
      color: var(--primary);
    }

    .courier-block {
      border-bottom: 1px solid var(--border);
    }

    .courier-block:last-child {
      border-bottom: none;
    }

    .courier-header {
      background: var(--light);
      padding: 1rem 2rem;
      font-weight: 600;
      color: var(--dark);
      border-bottom: 1px solid var(--border);
    }

    .week-block {
      margin: 0;
    }

    .toggle-btn {
      width: 100%;
      background: var(--white);
      border: none;
      padding: 1rem 2rem;
      text-align: left;
      font-size: 0.95rem;
      font-weight: 500;
      color: var(--secondary);
      cursor: pointer;
      transition: all 0.2s ease;
      border-bottom: 1px solid var(--border);
      display: flex;
      justify-content: space-between;
      align-items: center;
    }

    .toggle-btn:hover {
      background: var(--light);
      color: var(--dark);
    }

    .toggle-btn::after {
      content: '\f107';
      font-family: 'Font Awesome 6 Free';
      font-weight: 900;
      transition: transform 0.2s ease;
    }

    .toggle-btn.active::after {
      transform: rotate(180deg);
    }

    .week-content {
      padding: 0;
      overflow: hidden;
      transition: max-height 0.3s ease;
    }

    .data-table {
      width: 100%;
      border-collapse: collapse;
    }

    .data-table th {
      background: var(--light);
      padding: 1rem;
      text-align: left;
      font-weight: 600;
      color: var(--dark);
      font-size: 0.9rem;
      text-transform: uppercase;
      letter-spacing: 0.05em;
    }

    .data-table td {
      padding: 1rem;
      border-bottom: 1px solid var(--border);
      color: var(--secondary);
    }

    .data-table tr:hover td {
      background: var(--light);
    }

    .badge {
      display: inline-flex;
      align-items: center;
      gap: 0.25rem;
      padding: 0.25rem 0.75rem;
      border-radius: 999px;
      font-size: 0.8rem;
      font-weight: 500;
    }

    .badge.success {
      background: #dcfce7;
      color: #166534;
    }

    .badge.warning {
      background: #fef3c7;
      color: #92400e;
    }

    .badge.info {
      background: #dbeafe;
      color: #1e40af;
    }

    .metric {
      font-size: 1.5rem;
      font-weight: 700;
      color: var(--primary);
    }

    .no-data {
      text-align: center;
      padding: 3rem;
      color: var(--secondary);
    }

    .no-data i {
      font-size: 3rem;
      margin-bottom: 1rem;
      opacity: 0.5;
    }

    @media (max-width: 768px) {
      .container {
        padding: 1rem 0.5rem;
      }

      .header {
        padding: 1.5rem;
        flex-direction: column;
        text-align: center;
      }

      .header h1 {
        font-size: 1.5rem;
      }

      .stats-grid {
        grid-template-columns: 1fr;
      }

      .section-header {
        padding: 1rem 1.5rem;
      }

      .courier-header, .toggle-btn {
        padding: 0.75rem 1.5rem;
      }

      .data-table {
        font-size: 0.85rem;
      }

      .data-table th, .data-table td {
        padding: 0.75rem 0.5rem;
      }
    }

    /* Animation for expanding content */
    @keyframes slideDown {
      from {
        opacity: 0;
        transform: translateY(-10px);
      }
      to {
        opacity: 1;
        transform: translateY(0);
      }
    }

    .week-content[style*="block"] {
      animation: slideDown 0.3s ease;
    }
  </style>
</head>
<body>
  <div class="container">
    <!-- Header -->
    <div class="header">
      <h1>
        <i class="fas fa-chart-line"></i>
        Admin Dashboard
      </h1>
      <form action="/admin/export_excel" method="get" style="margin: 0;">
        <input type="hidden" name="init_data" id="export-init-data">
        <button type="submit" class="export-btn">
          <i class="fas fa-download"></i>
          Excel Export
        </button>
      </form>
    </div>

    <!-- Mock Statistics Cards -->
    <div class="stats-grid">
      <div class="stat-card">
        <div class="stat-header">
          <div class="stat-icon courier">
            <i class="fas fa-motorcycle"></i>
          </div>
          <div class="stat-title">Aktív Futárok</div>
        </div>
        <div class="metric" id="active-couriers">0</div>
        <p style="color: var(--secondary); font-size: 0.9rem; margin-top: 0.5rem;">
          <i class="fas fa-arrow-up" style="color: var(--success);"></i>
          +12% az elmúlt héten
        </p>
      </div>
      
      <div class="stat-card">
        <div class="stat-header">
          <div class="stat-icon restaurant">
            <i class="fas fa-utensils"></i>
          </div>
          <div class="stat-title">Éttermek</div>
        </div>
        <div class="metric" id="restaurants">0</div>
        <p style="color: var(--secondary); font-size: 0.9rem; margin-top: 0.5rem;">
          <i class="fas fa-arrow-up" style="color: var(--success);"></i>
          +3 új regisztráció
        </p>
      </div>
      
      <div class="stat-card">
        <div class="stat-header">
          <div class="stat-icon delivery">
            <i class="fas fa-box"></i>
          </div>
          <div class="stat-title">Kiszállítások</div>
        </div>
        <div class="metric" id="deliveries">0</div>
        <p style="color: var(--secondary); font-size: 0.9rem; margin-top: 0.5rem;">
          <i class="fas fa-clock" style="color: var(--warning);"></i>
          <span id="avg-time">0</span> perc átlag
        </p>
      </div>
    </div>

    <!-- Courier Statistics Section -->
    <div class="section">
      <div class="section-header">
        <div class="section-title">
          <i class="fas fa-motorcycle"></i>
          Heti Futár Bontás
        </div>
      </div>
      <div id="courier-stats">
        {% for courier, weeks in courier_stats.items() %}
        <div class="courier-block">
          <div class="courier-header">
            <i class="fas fa-user"></i>
            {{ courier }}
          </div>
          {% for r in weeks %}
            {% set weeknum = r.week.split('-')[1] %}
            {% set year = r.week.split('-')[0] %}
            
            <div class="week-block">
              <button class="toggle-btn" onclick="toggleWeek(this)">
                <span>{{ year }}. év, {{ weeknum }}. hét</span>
                <span class="badge info">{{ r.cnt }} rendelés</span>
              </button>
              <div class="week-content" style="display: none;">
                <table class="data-table">
                  <thead>
                    <tr>
                      <th><i class="fas fa-calendar"></i> Hét</th>
                      <th><i class="fas fa-box"></i> Darab</th>
                      <th><i class="fas fa-clock"></i> Átlag Idő (perc)</th>
                    </tr>
                  </thead>
                  <tbody>
                    <tr>
                      <td>{{ r.week }}</td>
                      <td><span class="badge success">{{ r.cnt }}</span></td>
                      <td>{{ r.avg_min }}</td>
                    </tr>
                  </tbody>
                </table>
              </div>
            </div>
          {% endfor %}
        </div>
        {% endfor %}
      </div>
    </div>

    <!-- Restaurant Statistics Section -->
    <div class="section">
      <div class="section-header">
        <div class="section-title">
          <i class="fas fa-utensils"></i>
          Étterem Bontás
        </div>
      </div>
      {% if weekly_restaurant %}
      <table class="data-table">
        <thead>
          <tr>
            <th><i class="fas fa-calendar"></i> Hét</th>
            <th><i class="fas fa-store"></i> Csoport</th>
            <th><i class="fas fa-box"></i> Darab</th>
            <th><i class="fas fa-clock"></i> Átlag Idő</th>
          </tr>
        </thead>
        <tbody>
          {% for r in weekly_restaurant %}
          <tr>
            <td>{{ r.week }}</td>
            <td>{{ r.group_name }}</td>
            <td><span class="badge info">{{ r.cnt }}</span></td>
            <td>{{ r.avg_min }} perc</td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
      {% else %}
      <div class="no-data">
        <i class="fas fa-chart-bar"></i>
        <p>Nincs elérhető adat</p>
      </div>
      {% endif %}
    </div>

    <!-- Detailed Deliveries Section -->
    <div class="section">
      <div class="section-header">
        <div class="section-title">
          <i class="fas fa-shipping-fast"></i>
          Részletes Kézbesítések
        </div>
      </div>
      {% if deliveries %}
      <div style="overflow-x: auto;">
        <table class="data-table">
          <thead>
            <tr>
              <th><i class="fas fa-calendar"></i> Dátum</th>
              <th><i class="fas fa-user"></i> Futár</th>
              <th><i class="fas fa-store"></i> Csoport</th>
              <th><i class="fas fa-map-marker"></i> Cím</th>
              <th><i class="fas fa-clock"></i> Idő (perc)</th>
            </tr>
          </thead>
          <tbody>
            {% for r in deliveries %}
            <tr>
              <td>{{ r.delivered_at }}</td>
              <td>{{ r.courier_name or r.delivery_partner_id }}</td>
              <td>{{ r.group_name }}</td>
              <td style="max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">
                {{ r.restaurant_address }}
              </td>
              <td>
                {% if r.min %}
                  <span class="badge {% if r.min|float < 20 %}success{% elif r.min|float < 40 %}warning{% else %}info{% endif %}">
                    {{ r.min }}
                  </span>
                {% else %}
                  <span class="badge">N/A</span>
                {% endif %}
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      {% else %}
      <div class="no-data">
        <i class="fas fa-truck"></i>
        <p>Nincs kézbesítési adat</p>
      </div>
      {% endif %}
    </div>
  </div>

  <script>
    // Set export form data
    document.getElementById("export-init-data").value = window.Telegram?.WebApp?.initData || '';

    // Toggle week details
    function toggleWeek(btn) {
      const content = btn.nextElementSibling;
      const isVisible = content.style.display === "block";
      
      if (isVisible) {
        content.style.display = "none";
        btn.classList.remove('active');
      } else {
        content.style.display = "block";
        btn.classList.add('active');
      }
    }

    // Mock data for statistics (replace with actual data from your backend)
    function updateStats() {
      // These would be populated from your actual data
      const courierStats = {{ courier_stats|tojson }};
      const deliveryStats = {{ deliveries|tojson }};
      const restaurantStats = {{ weekly_restaurant|tojson }};

      // Count unique couriers
      const uniqueCouriers = Object.keys(courierStats || {}).length;
      document.getElementById('active-couriers').textContent = uniqueCouriers;

      // Count unique restaurants
      const uniqueRestaurants = new Set((restaurantStats || []).map(r => r.group_name)).size;
      document.getElementById('restaurants').textContent = uniqueRestaurants;

      // Total deliveries and average time
      const totalDeliveries = (deliveryStats || []).length;
      document.getElementById('deliveries').textContent = totalDeliveries;

      if (deliveryStats && deliveryStats.length > 0) {
        const avgTime = deliveryStats
          .filter(d => d.min)
          .reduce((sum, d) => sum + parseFloat(d.min), 0) / 
          deliveryStats.filter(d => d.min).length;
        document.getElementById('avg-time').textContent = Math.round(avgTime);
      }
    }

    // Delete order function for admin
    async function deleteOrder(orderId) {
      if (!confirm(`Biztosan törölni szeretnéd a #${orderId} rendelést?`)) {
        return;
      }
      
      try {
        const response = await fetch('/api/delete_order', {
          method: 'POST',
          headers: {'Content-Type': 'application/json'},
          body: JSON.stringify({
            initData: window.Telegram?.WebApp?.initData || '',
            order_id: orderId
          })
        });
        
        const data = await response.json();
        if (data.ok) {
          alert('Rendelés sikeresen törölve!');
          location.reload(); // Refresh the page to show updated data
        } else {
          alert('Hiba a törlés során: ' + (data.error || 'Ismeretlen hiba'));
        }
      } catch (e) {
        alert('Hálózati hiba a törlés során: ' + e.message);
      }
    }

    // Initialize stats
    updateStats();

    // Add smooth scrolling
    document.querySelectorAll('a[href^="#"]').forEach(anchor => {
      anchor.addEventListener('click', function (e) {
        e.preventDefault();
        document.querySelector(this.getAttribute('href')).scrollIntoView({
          behavior: 'smooth'
        });
      });
    });
  </script>
</body>
</html>
"""

# --- other endpoints preserved (accept/pickup/mark_delivered etc) ---
@app.route("/api/orders")
def api_orders():
    try: return jsonify(db.get_open_orders())
    except Exception as e: logger.error(f"api_orders error: {e}"); return jsonify([])

@app.route("/api/accept_order", methods=["POST"])
def api_accept_order():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        eta = int(data.get("estimated_time", 20))
        user = validate_telegram_data(data.get("initData", ""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        order = db.get_order_by_id(order_id)
        if not order or order["status"] != "pending": return jsonify({"ok": False, "error": "not_available"}), 400
        partner_name = ((user.get("first_name","") + " " + user.get("last_name","")).strip()) or str(user.get("id"))
        partner_username = user.get("username")
        db.update_order_status(order_id, "accepted", partner_id=user.get("id"), partner_name=partner_name, partner_username=partner_username, estimated_time=eta)
        try:
            partner_contact = f"@{partner_username}" if partner_username else partner_name
            text = ("🚚 **FUTÁR JELENTKEZETT!**\n\n" f"👤 **Futár:** {partner_name}\n" f"📱 **Kontakt:** {partner_contact}\n" f"⏱️ **Becsült érkezés:** {eta} perc\n" f"📋 **Rendelés ID:** #{order_id}\n")
            notification_queue.put({"chat_id": order["group_id"], "text": text})
        except Exception as e: logger.error(f"group notify fail (accept): {e}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_accept_order error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/pickup_order", methods=["POST"])
def api_pickup_order():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        order = db.get_order_by_id(order_id)
        if not order or order["status"] != "accepted" or order.get("delivery_partner_id") not in (None, user.get("id")):
            pass
        db.update_order_status(order_id, "picked_up", partner_id=user.get("id"))
        try:
            partner_name = ((user.get("first_name","") + " " + user.get("last_name","")).strip()) or str(user.get("id"))
            partner_username = user.get("username"); partner_contact = f"@{partner_username}" if partner_username else partner_name
            text = ("📦 **RENDELÉS FELVÉVE!**\n\n" f"👤 **Futár:** {partner_name}\n" f"📱 **Kontakt:** {partner_contact}\n" f"📋 **Rendelés ID:** #{order_id}\n")
            notification_queue.put({"chat_id": order["group_id"], "text": text})
        except Exception as e: logger.error(f"group notify fail (pickup): {e}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_pickup_order error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/mark_delivered", methods=["POST"])
def api_mark_delivered():
    try:
        data = request.json or {}
        order_id = int(data.get("order_id"))
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error": "unauthorized"}), 401
        order = db.get_order_by_id(order_id)
        if not order or order["status"] != "picked_up": return jsonify({"ok": False, "error": "not_pickup"}), 400
        db.update_order_status(order_id, "delivered")
        try:
            partner_name = ((user.get("first_name","") + " " + user.get("last_name","")).strip()) or str(user.get("id"))
            partner_username = user.get("username"); partner_contact = f"@{partner_username}" if partner_username else partner_name
            text = ("✅ **RENDELÉS KISZÁLLÍTVA!**\n\n" f"👤 **Futár:** {partner_name}\n" f"📱 **Kontakt:** {partner_contact}\n" f"📋 **Rendelés ID:** #{order_id}\n")
            notification_queue.put({"chat_id": order["group_id"], "text": text})
        except Exception as e: logger.error(f"group notify fail (delivered): {e}")
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_mark_delivered error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/orders_by_status", methods=["GET"])
def api_orders_by_status():
    try:
        status = (request.args.get("status") or "").strip()
        courier_id = request.args.get("courier_id", type=int)
        if not status: return jsonify([])
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        if status == "pending":
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status FROM orders WHERE status='pending' AND DATE(created_at) = DATE('now', 'localtime') ORDER BY created_at DESC")
            rows = [dict(r) for r in cur.fetchall()]
        elif status in ("accepted","picked_up","delivered"):
            if not courier_id: conn.close(); return jsonify({"ok": False, "error": "missing_courier"}), 400
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, estimated_time FROM orders WHERE status=? AND delivery_partner_id=? ORDER BY created_at DESC", (status, courier_id))
            rows = [dict(r) for r in cur.fetchall()]
        else:
            rows = []
        conn.close(); return jsonify(rows)
    except Exception as e:
        logger.error(f"api_orders_by_status error: {e}"); return jsonify([]), 500

@app.route("/api/my_orders", methods=["POST"])
def api_my_orders():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error":"unauthorized"}), 401
        status = data.get("status","").strip()
        if status not in ("accepted","picked_up","delivered"): return jsonify({"ok": True, "orders": []})
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        if status == "delivered":
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, estimated_time FROM orders WHERE status=? AND delivery_partner_id=? AND DATE(created_at) = DATE('now', 'localtime') ORDER BY created_at DESC", (status, user["id"]))
        else:
            cur.execute("SELECT id, restaurant_name, restaurant_address, phone_number, order_details, group_id, group_name, created_at, status, estimated_time FROM orders WHERE status=? AND delivery_partner_id=? ORDER BY created_at DESC", (status, user["id"]))
        rows = [dict(r) for r in cur.fetchall()]; conn.close(); return jsonify({"ok": True, "orders": rows})
    except Exception as e:
        logger.error(f"api_my_orders error: {e}"); return jsonify({"ok": False, "error": str(e)}), 500
    
@app.route("/admin")
def admin_page():
    init_data = request.args.get('init_data', ''); user = validate_telegram_data(init_data)
    if not user or user.get("id") not in ADMIN_USER_IDS: return "🚫 Hozzáférés megtagadva", 403
    try:
        import datetime
        current_week = datetime.date.today().strftime("%Y-%W")
        conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()
        cur.execute("""
            SELECT strftime('%Y-%W', delivered_at) AS week, delivery_partner_id, COALESCE(delivery_partner_name, '') AS courier_name, COUNT(*) AS cnt, ROUND(AVG((julianday(delivered_at) - julianday(accepted_at)) * 24 * 60), 1) AS avg_min FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL GROUP BY delivery_partner_id, week ORDER BY week DESC, cnt DESC
        """)
        weekly_courier = [dict(r) for r in cur.fetchall()]
        from collections import defaultdict
        courier_stats = defaultdict(list)
        for r in weekly_courier:
            courier_stats[r["courier_name"] or r["delivery_partner_id"]].append(r)
        cur.execute("""
            SELECT strftime('%Y-%W', delivered_at) AS week, group_name, COUNT(*) AS cnt, ROUND(AVG((julianday(delivered_at) - julianday(accepted_at)) * 24 * 60), 1) AS avg_min FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL GROUP BY group_name, week ORDER BY week DESC, cnt DESC
        """)
        weekly_restaurant = [dict(r) for r in cur.fetchall()]
        cur.execute("""
            SELECT delivered_at, delivery_partner_id, COALESCE(delivery_partner_name, '') AS courier_name, group_name, restaurant_address, ROUND((julianday(delivered_at) - julianday(accepted_at)) * 24 * 60, 1) AS min FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL ORDER BY delivered_at DESC LIMIT 500
        """)
        deliveries = [dict(r) for r in cur.fetchall()]; conn.close()
        return render_template_string(ADMIN_HTML,
                                      courier_stats=courier_stats,
                                      weekly_restaurant=weekly_restaurant,
                                      deliveries=deliveries,
                                      current_week=current_week)
    except Exception as e:
        logger.error(f"admin_page error: {e}"); return "admin error", 500

@app.route("/api/is_admin", methods=["POST"])
def api_is_admin():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData", ""))
        if not user: return jsonify({"ok": False, "admin": False}), 401
        return jsonify({"ok": True, "admin": user.get("id") in ADMIN_USER_IDS})
    except Exception as e:
        logger.error(f"api_is_admin error: {e}"); return jsonify({"ok": False, "admin": False}), 500

@app.route("/admin/export_excel")
def admin_export_excel():

    conn = sqlite3.connect(DB_NAME); conn.row_factory = sqlite3.Row; cur = conn.cursor()

    # ugyanaz a 3 lekérdezés, mint admin_page-ben
    cur.execute("""SELECT strftime('%Y-%W', delivered_at) AS week, delivery_partner_id, 
                   COALESCE(delivery_partner_name, '') AS courier_name, COUNT(*) AS cnt,
                   ROUND(AVG((julianday(delivered_at)-julianday(accepted_at))*24*60),1) AS avg_min
                   FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL
                   GROUP BY delivery_partner_id, week ORDER BY week DESC, cnt DESC""")
    weekly_courier = [dict(r) for r in cur.fetchall()]

    cur.execute("""SELECT strftime('%Y-%W', delivered_at) AS week, group_name, COUNT(*) AS cnt,
                   ROUND(AVG((julianday(delivered_at)-julianday(accepted_at))*24*60),1) AS avg_min
                   FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL
                   GROUP BY group_name, week ORDER BY week DESC, cnt DESC""")
    weekly_restaurant = [dict(r) for r in cur.fetchall()]

    cur.execute("""SELECT delivered_at, delivery_partner_id, COALESCE(delivery_partner_name, '') AS courier_name,
                   group_name, restaurant_address,
                   ROUND((julianday(delivered_at)-julianday(accepted_at))*24*60,1) AS min
                   FROM orders WHERE delivered_at IS NOT NULL AND accepted_at IS NOT NULL
                   ORDER BY delivered_at DESC LIMIT 500""")
    deliveries = [dict(r) for r in cur.fetchall()]
    conn.close()

    # Excel létrehozás
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Heti futár"
    ws1.append(["Hét", "Futár", "Darab", "Átlag idő (perc)"])
    for r in weekly_courier:
        ws1.append([r["week"], r["courier_name"] or r["delivery_partner_id"], r["cnt"], r["avg_min"]])

    ws2 = wb.create_sheet("Heti étterem")
    ws2.append(["Hét", "Csoport", "Darab", "Átlag idő (perc)"])
    for r in weekly_restaurant:
        ws2.append([r["week"], r["group_name"], r["cnt"], r["avg_min"]])

    # Éttermek szétválogatása
    deliveries_by_restaurant = {}
    for r in deliveries:
        gname = r["group_name"] or "Ismeretlen"
        deliveries_by_restaurant.setdefault(gname, []).append(r)

    # Minden étteremhez külön munkalap
    for gname, rows in deliveries_by_restaurant.items():
       # Excel sheet neve max 31 karakter lehet → levágjuk, ha hosszabb
        safe_name = gname[:31]
        ws = wb.create_sheet(safe_name)
        ws.append(["Dátum", "Futár", "Csoport", "Cím", "Idő (perc)"])
        for r in rows:
            ws.append([r["delivered_at"], r["courier_name"] or r["delivery_partner_id"],
                       r["group_name"], r["restaurant_address"], r["min"]])
    # mentés memóriába
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name="statisztika.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/my_orders_history", methods=["POST"])
def api_my_orders_history():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData",""))
        if not user: return jsonify({"ok": False, "error":"unauthorized"}), 401
        
        week = data.get("week", "").strip()  # formátum: "2025-01"
        
        conn = sqlite3.connect(DB_NAME)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        if week:
            cur.execute("""
                SELECT id, restaurant_name, restaurant_address, phone_number, order_details, 
                       group_id, group_name, created_at, status, estimated_time, delivered_at
                FROM orders 
                WHERE delivery_partner_id=? AND strftime('%Y-%W', created_at) = ?
                ORDER BY created_at DESC
            """, (user["id"], week))
        else:
            # Összes hét listázása
            cur.execute("""
                SELECT DISTINCT strftime('%Y-%W', created_at) as week, 
                       COUNT(*) as count
                FROM orders 
                WHERE delivery_partner_id=?
                GROUP BY strftime('%Y-%W', created_at)
                ORDER BY week DESC
            """, (user["id"],))
        
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return jsonify({"ok": True, "data": rows})
    except Exception as e:
        logger.error(f"api_my_orders_history error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/register_courier", methods=["POST"])
def api_register_courier():
    try:
        data = request.json or {}
        user = validate_telegram_data(data.get("initData",""))
        if not user:
            return jsonify({"ok": False, "error": "unauthorized"}), 401
        db.register_courier(user)
        return jsonify({"ok": True})
    except Exception as e:
        logger.error(f"api_register_courier error: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ---------------- Server start ----------------
def run_flask():
    from werkzeug.serving import WSGIRequestHandler
    WSGIRequestHandler.protocol_version = "HTTP/1.1"
    app.run(host="0.0.0.0", port=5000, debug=False)

if __name__ == "__main__":
    # start flask in background thread and start bot polling (if available)
    threading.Thread(target=run_flask, daemon=True).start()
    if TELEGRAM_AVAILABLE:
        RestaurantBot().run()
    else:
        logger.info("Flask started in this environment; Telegram bot not available.")
